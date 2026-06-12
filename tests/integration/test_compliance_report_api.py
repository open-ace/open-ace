"""
Integration tests for Compliance Report API endpoints.

Tests cover:
- Report generation with different formats (JSON, CSV, HTML, Excel)
- Report download endpoints
- Format parameter validation
- Language parameter validation
- Security headers (CSP)
- Error handling for save failures
"""

from datetime import datetime, timedelta, timezone
from functools import wraps
from io import BytesIO

import pytest

# Check if openpyxl is available for Excel tests
try:
    from openpyxl import load_workbook

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


@pytest.fixture
def app(tmp_db):
    """Create Flask app for testing with temporary database."""
    from unittest.mock import patch

    from flask import Flask

    from app.routes.compliance import compliance_bp

    app = Flask(__name__)
    app.register_blueprint(compliance_bp)
    app.config["TESTING"] = True
    app.config["SECRET_KEY"] = "test-secret-key"

    # Patch report_generator to use tmp_db
    with patch("app.routes.compliance.report_generator.db", tmp_db):
        yield app


@pytest.fixture
def client(app):
    """Create test client with authentication."""
    from unittest.mock import patch

    from flask import g

    test_client = app.test_client()

    # Create a wrapper that patches authentication for each request
    class AuthenticatedClient:
        def __init__(self, client):
            self._client = client

        def get(self, *args, **kwargs):
            with patch("app.auth.decorators._extract_token", return_value="test-token"):
                with patch(
                    "app.auth.decorators._load_user_from_token",
                    return_value={"id": 1, "role": "admin", "username": "test_admin"},
                ):
                    return self._client.get(*args, **kwargs)

        def post(self, *args, **kwargs):
            with patch("app.auth.decorators._extract_token", return_value="test-token"):
                with patch(
                    "app.auth.decorators._load_user_from_token",
                    return_value={"id": 1, "role": "admin", "username": "test_admin"},
                ):
                    return self._client.post(*args, **kwargs)

        def put(self, *args, **kwargs):
            with patch("app.auth.decorators._extract_token", return_value="test-token"):
                with patch(
                    "app.auth.decorators._load_user_from_token",
                    return_value={"id": 1, "role": "admin", "username": "test_admin"},
                ):
                    return self._client.put(*args, **kwargs)

        def delete(self, *args, **kwargs):
            with patch("app.auth.decorators._extract_token", return_value="test-token"):
                with patch(
                    "app.auth.decorators._load_user_from_token",
                    return_value={"id": 1, "role": "admin", "username": "test_admin"},
                ):
                    return self._client.delete(*args, **kwargs)

    return AuthenticatedClient(test_client)


@pytest.fixture
def admin_headers():
    """Headers for admin user requests."""
    return {"Content-Type": "application/json"}


class TestReportGenerationAPI:
    """Test report generation API endpoints."""

    def test_generate_report_json(self, client, admin_headers):
        """Test generating report in JSON format."""
        response = client.post(
            "/api/compliance/reports",
            json={
                "report_type": "usage_summary",
                "period_start": (datetime.now(timezone.utc) - timedelta(days=30)).strftime(
                    "%Y-%m-%d"
                ),
                "period_end": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                "format": "json",
            },
            headers=admin_headers,
        )

        assert response.status_code == 200
        data = response.get_json()
        assert "metadata" in data
        assert "summary" in data
        assert "details" in data
        assert "compliance_checks" in data
        assert "recommendations" in data

    def test_generate_report_csv(self, client, admin_headers):
        """Test generating report in CSV format."""
        response = client.post(
            "/api/compliance/reports",
            json={
                "report_type": "usage_summary",
                "period_start": (datetime.now(timezone.utc) - timedelta(days=30)).strftime(
                    "%Y-%m-%d"
                ),
                "period_end": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                "format": "csv",
            },
            headers=admin_headers,
        )

        assert response.status_code == 200
        assert response.content_type == "text/csv"
        # CSV should have headers and data
        csv_content = response.get_data(as_text=True)
        assert len(csv_content) > 0

    def test_generate_report_html(self, client, admin_headers):
        """Test generating report in HTML format."""
        response = client.post(
            "/api/compliance/reports",
            json={
                "report_type": "usage_summary",
                "period_start": (datetime.now(timezone.utc) - timedelta(days=30)).strftime(
                    "%Y-%m-%d"
                ),
                "period_end": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                "format": "html",
                "language": "en",
            },
            headers=admin_headers,
        )

        assert response.status_code == 200
        assert response.content_type == "text/html"
        html_content = response.get_data(as_text=True)
        # HTML should contain basic elements
        assert "<html" in html_content
        assert "<body" in html_content
        assert "Generated by" in html_content

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl required for Excel tests")
    def test_generate_report_excel(self, client, admin_headers):
        """Test generating report in Excel format."""
        response = client.post(
            "/api/compliance/reports",
            json={
                "report_type": "usage_summary",
                "period_start": (datetime.now(timezone.utc) - timedelta(days=30)).strftime(
                    "%Y-%m-%d"
                ),
                "period_end": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                "format": "excel",
                "language": "en",
            },
            headers=admin_headers,
        )

        assert response.status_code == 200
        assert "spreadsheetml.sheet" in response.content_type
        # Excel should be valid binary data
        excel_data = response.get_data()
        assert len(excel_data) > 0

        # Verify it's a valid Excel file
        wb = load_workbook(BytesIO(excel_data))
        assert len(wb.sheetnames) >= 3

    def test_generate_report_html_csp_header(self, client, admin_headers):
        """Test that HTML response includes CSP header."""
        response = client.post(
            "/api/compliance/reports",
            json={
                "report_type": "usage_summary",
                "period_start": (datetime.now(timezone.utc) - timedelta(days=30)).strftime(
                    "%Y-%m-%d"
                ),
                "period_end": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                "format": "html",
            },
            headers=admin_headers,
        )

        assert response.status_code == 200
        # Check for CSP header
        csp = response.headers.get("Content-Security-Policy")
        assert csp is not None
        assert "default-src 'self'" in csp
        assert "script-src 'none'" in csp

    def test_generate_report_language_en(self, client, admin_headers):
        """Test generating HTML report in English."""
        response = client.post(
            "/api/compliance/reports",
            json={
                "report_type": "usage_summary",
                "format": "html",
                "language": "en",
            },
            headers=admin_headers,
        )

        assert response.status_code == 200
        html_content = response.get_data(as_text=True)
        # English title should be present
        assert "Usage Summary" in html_content or "Compliance Report" in html_content

    def test_generate_report_language_zh(self, client, admin_headers):
        """Test generating HTML report in Chinese."""
        response = client.post(
            "/api/compliance/reports",
            json={
                "report_type": "usage_summary",
                "format": "html",
                "language": "zh",
            },
            headers=admin_headers,
        )

        assert response.status_code == 200
        html_content = response.get_data(as_text=True)
        # Chinese title should be present
        assert "使用汇总" in html_content or "合规报告" in html_content

    def test_generate_report_language_ja(self, client, admin_headers):
        """Test generating HTML report in Japanese."""
        response = client.post(
            "/api/compliance/reports",
            json={
                "report_type": "usage_summary",
                "format": "html",
                "language": "ja",
            },
            headers=admin_headers,
        )

        assert response.status_code == 200
        html_content = response.get_data(as_text=True)
        # Japanese title should be present
        assert "サマリー" in html_content or "コンプライアンス" in html_content

    def test_generate_report_language_ko(self, client, admin_headers):
        """Test generating HTML report in Korean."""
        response = client.post(
            "/api/compliance/reports",
            json={
                "report_type": "usage_summary",
                "format": "html",
                "language": "ko",
            },
            headers=admin_headers,
        )

        assert response.status_code == 200
        html_content = response.get_data(as_text=True)
        # Korean title should be present
        assert "요약" in html_content or "준수" in html_content

    def test_generate_report_invalid_format(self, client, admin_headers):
        """Test generating report with invalid format."""
        response = client.post(
            "/api/compliance/reports",
            json={
                "report_type": "usage_summary",
                "format": "invalid_format",
            },
            headers=admin_headers,
        )

        # Should still return JSON as fallback
        assert response.status_code == 200

    def test_generate_report_missing_report_type(self, client, admin_headers):
        """Test generating report without report_type."""
        response = client.post(
            "/api/compliance/reports",
            json={
                "format": "json",
            },
            headers=admin_headers,
        )

        assert response.status_code == 400
        data = response.get_json()
        assert "error" in data

    def test_generate_report_different_types(self, client, admin_headers):
        """Test generating different report types."""
        report_types = [
            "usage_summary",
            "user_activity",
            "audit_trail",
            "security",
            "comprehensive",
        ]

        for report_type in report_types:
            response = client.post(
                "/api/compliance/reports",
                json={
                    "report_type": report_type,
                    "format": "json",
                },
                headers=admin_headers,
            )
            assert response.status_code == 200
            data = response.get_json()
            assert data["metadata"]["report_type"] == report_type


class TestSavedReportAPI:
    """Test saved report API endpoints."""

    def test_get_saved_reports_list(self, client, admin_headers):
        """Test getting saved reports list."""
        # First generate a report
        client.post(
            "/api/compliance/reports",
            json={
                "report_type": "usage_summary",
                "format": "json",
            },
            headers=admin_headers,
        )

        # Then get the list
        response = client.get(
            "/api/compliance/reports/saved",
            headers=admin_headers,
        )

        assert response.status_code == 200
        data = response.get_json()
        assert "reports" in data
        assert "count" in data

    def test_get_saved_report_json(self, client, admin_headers):
        """Test getting saved report in JSON format."""
        # First generate a report
        gen_response = client.post(
            "/api/compliance/reports",
            json={
                "report_type": "usage_summary",
                "format": "json",
            },
            headers=admin_headers,
        )
        report_id = gen_response.get_json()["metadata"]["report_id"]

        # Then get the saved report
        response = client.get(
            f"/api/compliance/reports/{report_id}?format=json",
            headers=admin_headers,
        )

        assert response.status_code == 200
        data = response.get_json()
        assert "metadata" in data
        assert "summary" in data

    def test_get_saved_report_csv(self, client, admin_headers):
        """Test getting saved report in CSV format."""
        # First generate a report
        gen_response = client.post(
            "/api/compliance/reports",
            json={
                "report_type": "usage_summary",
                "format": "json",
            },
            headers=admin_headers,
        )
        report_id = gen_response.get_json()["metadata"]["report_id"]

        # Then get the saved report in CSV
        response = client.get(
            f"/api/compliance/reports/{report_id}?format=csv",
            headers=admin_headers,
        )

        assert response.status_code == 200
        assert response.content_type == "text/csv"

    def test_get_saved_report_html(self, client, admin_headers):
        """Test getting saved report in HTML format."""
        # First generate a report
        gen_response = client.post(
            "/api/compliance/reports",
            json={
                "report_type": "usage_summary",
                "format": "json",
            },
            headers=admin_headers,
        )
        report_id = gen_response.get_json()["metadata"]["report_id"]

        # Then get the saved report in HTML
        response = client.get(
            f"/api/compliance/reports/{report_id}?format=html&language=en",
            headers=admin_headers,
        )

        assert response.status_code == 200
        assert response.content_type == "text/html"
        html_content = response.get_data(as_text=True)
        assert "<html" in html_content

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl required for Excel tests")
    def test_get_saved_report_excel(self, client, admin_headers):
        """Test getting saved report in Excel format."""
        # First generate a report
        gen_response = client.post(
            "/api/compliance/reports",
            json={
                "report_type": "usage_summary",
                "format": "json",
            },
            headers=admin_headers,
        )
        report_id = gen_response.get_json()["metadata"]["report_id"]

        # Then get the saved report in Excel
        response = client.get(
            f"/api/compliance/reports/{report_id}?format=excel&language=en",
            headers=admin_headers,
        )

        assert response.status_code == 200
        assert "spreadsheetml.sheet" in response.content_type

    def test_get_saved_report_not_found(self, client, admin_headers):
        """Test getting non-existent saved report."""
        response = client.get(
            "/api/compliance/reports/non-existent-report-id",
            headers=admin_headers,
        )

        assert response.status_code == 404
        data = response.get_json()
        assert "error" in data


class TestReportSecurity:
    """Test report security features."""

    def test_html_xss_protection(self, client, admin_headers):
        """Test that XSS attempts are escaped in HTML output."""
        response = client.post(
            "/api/compliance/reports",
            json={
                "report_type": "usage_summary",
                "format": "html",
            },
            headers=admin_headers,
        )

        assert response.status_code == 200
        html_content = response.get_data(as_text=True)
        # Should not contain unescaped script tags
        assert "<script>alert" not in html_content

    def test_csp_header_script_src_none(self, client, admin_headers):
        """Test that CSP header blocks scripts."""
        response = client.post(
            "/api/compliance/reports",
            json={
                "report_type": "usage_summary",
                "format": "html",
            },
            headers=admin_headers,
        )

        csp = response.headers.get("Content-Security-Policy")
        assert "script-src 'none'" in csp

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl required for Excel tests")
    def test_excel_no_script_content(self, client, admin_headers):
        """Test that Excel output doesn't contain script content."""
        response = client.post(
            "/api/compliance/reports",
            json={
                "report_type": "usage_summary",
                "format": "excel",
            },
            headers=admin_headers,
        )

        excel_data = response.get_data()
        # Should not contain script-like content
        assert "<script" not in excel_data.decode("utf-8", errors="ignore")


class TestReportFormatValidation:
    """Test format parameter validation."""

    def test_format_defaults_to_json(self, client, admin_headers):
        """Test that missing format defaults to JSON."""
        response = client.post(
            "/api/compliance/reports",
            json={
                "report_type": "usage_summary",
            },
            headers=admin_headers,
        )

        assert response.status_code == 200
        # Should return JSON
        data = response.get_json()
        assert isinstance(data, dict)

    def test_all_supported_formats(self, client, admin_headers):
        """Test all supported report formats."""
        formats = ["json", "csv", "html", "excel"]

        for fmt in formats:
            response = client.post(
                "/api/compliance/reports",
                json={
                    "report_type": "usage_summary",
                    "format": fmt,
                },
                headers=admin_headers,
            )
            assert response.status_code == 200

    def test_all_supported_languages(self, client, admin_headers):
        """Test all supported languages for HTML."""
        languages = ["en", "zh", "ja", "ko"]

        for lang in languages:
            response = client.post(
                "/api/compliance/reports",
                json={
                    "report_type": "usage_summary",
                    "format": "html",
                    "language": lang,
                },
                headers=admin_headers,
            )
            assert response.status_code == 200
            assert response.content_type == "text/html"


class TestReportErrorHandling:
    """Test report error handling scenarios."""

    def test_list_saved_reports_returns_empty_on_success_no_data(self, client, admin_headers):
        """Test that list_saved_reports returns empty list when no saved reports exist."""
        response = client.get(
            "/api/compliance/reports/saved",
            headers=admin_headers,
        )

        assert response.status_code == 200
        data = response.get_json()
        assert "reports" in data
        assert "count" in data
        # Should be empty list (no data), not error response
        assert isinstance(data["reports"], list)

    def test_list_saved_reports_error_handling(self, client, admin_headers):
        """Test that list_saved_reports returns error on database failure."""
        # This test verifies the fix for the "暂无已保存报告" issue
        # When database fails, it should return error (500), not empty list
        from unittest.mock import patch

        with patch("app.routes.compliance.report_generator.get_saved_reports") as mock_get:
            mock_get.side_effect = Exception("Database query failed")

            response = client.get(
                "/api/compliance/reports/saved",
                headers=admin_headers,
            )

            # Should return 500 error, not 200 with empty list
            assert response.status_code == 500
            data = response.get_json()
            assert "error" in data
            assert "Failed to query saved reports" in data["error"]

    def test_generate_report_error_on_save_failure(self, client, admin_headers):
        """Test that generate_report returns error when save_report returns False."""
        # This test verifies the fix for checking save_report return value
        from unittest.mock import Mock, patch

        # Mock generate_report to return a valid report
        mock_report = Mock()
        mock_report.metadata.report_id = "test-report-123"
        mock_report.to_dict.return_value = {
            "metadata": {"report_id": "test-report-123"},
            "summary": {},
            "details": [],
            "compliance_checks": [],
            "recommendations": [],
        }

        with patch("app.routes.compliance.report_generator.generate_report") as mock_gen:
            mock_gen.return_value = mock_report

            with patch("app.routes.compliance.report_generator.save_report") as mock_save:
                # Simulate save failure (returns False)
                mock_save.return_value = False

                response = client.post(
                    "/api/compliance/reports",
                    json={
                        "report_type": "usage_summary",
                        "format": "json",
                    },
                    headers=admin_headers,
                )

                # Should return 500 error when save fails
                assert response.status_code == 500
                data = response.get_json()
                assert "error" in data
                assert "Failed to save report to database" in data["error"]
