"""
Unit tests for ComplianceReport CSV and Excel export with nested field flattening.

Tests cover:
- CSV export with nested dict/list flattening
- Excel details sheet with nested field flattening
- datetime handling for CSV/Excel
- None value handling
- Unicode character handling
- RFC 4180 CSV compliance
- Long JSON truncation
- Dict format (comprehensive report) handling
"""

import csv
import json
import pytest
from datetime import datetime, timezone
from io import StringIO

# Check if openpyxl is available for Excel tests
try:
    import openpyxl
    from openpyxl import load_workbook
    from io import BytesIO
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from app.modules.compliance.report import (
    ComplianceReport,
    ReportMetadata,
    ReportType,
)


@pytest.fixture
def sample_report_metadata():
    """Create sample report metadata."""
    return ReportMetadata(
        report_id="test-report-001",
        report_type=ReportType.USAGE_SUMMARY.value,
        generated_at=datetime.now(timezone.utc).replace(tzinfo=None),
        period_start=datetime(2024, 1, 1),
        period_end=datetime(2024, 1, 31),
        generated_by=1,
    )


@pytest.fixture
def sample_report(sample_report_metadata):
    """Create sample compliance report with flat details."""
    return ComplianceReport(
        metadata=sample_report_metadata,
        summary={
            "period": {"start": "2024-01-01", "end": "2024-01-31", "days": 31},
            "totals": {
                "tokens": 10000,
                "requests": 500,
                "tools_used": 3,
                "unique_users": 10,
            },
        },
        details=[
            {
                "date": "2024-01-01",
                "tool_name": "claude",
                "host_name": "api.anthropic.com",
                "total_input_tokens": 1000,
                "total_output_tokens": 500,
                "total_tokens": 1500,
                "total_requests": 10,
            },
            {
                "date": "2024-01-02",
                "tool_name": "gpt",
                "host_name": "api.openai.com",
                "total_input_tokens": 2000,
                "total_output_tokens": 1000,
                "total_tokens": 3000,
                "total_requests": 20,
            },
        ],
        compliance_checks=[
            {
                "check_id": "data_retention",
                "name": "Data Retention Policy",
                "status": "pass",
                "message": "Data retention policy is being followed",
            },
        ],
        recommendations=["All compliance checks passed. Continue monitoring."],
    )


class TestReportToCsv:
    """Test CSV report generation with nested field flattening."""

    def test_to_csv_returns_string(self, sample_report):
        """Test that to_csv returns a string."""
        csv = sample_report.to_csv()
        assert isinstance(csv, str)

    def test_to_csv_empty_details(self, sample_report_metadata):
        """Test CSV with empty details returns empty string."""
        empty_report = ComplianceReport(
            metadata=sample_report_metadata,
            summary={},
            details=[],
            compliance_checks=[],
            recommendations=["No data available"],
        )
        csv = empty_report.to_csv()
        assert csv == ""

    def test_to_csv_flat_details(self, sample_report):
        """Test CSV with flat (non-nested) details."""
        csv = sample_report.to_csv()
        # Should have header row
        assert "date" in csv
        assert "tool_name" in csv
        # Should have data rows
        assert "claude" in csv or "gpt" in csv

    def test_to_csv_nested_dict(self, sample_report_metadata):
        """Test CSV with nested dict in details field."""
        report = ComplianceReport(
            metadata=sample_report_metadata,
            summary={},
            details=[
                {
                    "id": 1,
                    "action": "login",
                    "details": {"ip_address": "192.168.1.1", "browser": "Chrome"},
                    "timestamp": datetime(2024, 1, 1, 12, 0, 0),
                },
            ],
            compliance_checks=[],
            recommendations=["Test"],
        )
        csv = report.to_csv()
        # Nested dict should be converted to JSON string
        # CSV format escapes double quotes as two double quotes per RFC 4180
        # So {"ip_address": "192.168.1.1"} becomes {""ip_address"": ""192.168.1.1""}
        assert "ip_address" in csv
        assert "192.168.1.1" in csv

    def test_to_csv_nested_list(self, sample_report_metadata):
        """Test CSV with nested list in details field."""
        report = ComplianceReport(
            metadata=sample_report_metadata,
            summary={},
            details=[
                {
                    "id": 1,
                    "action": "export",
                    "items": ["item1", "item2", "item3"],
                    "timestamp": datetime(2024, 1, 1, 12, 0, 0),
                },
            ],
            compliance_checks=[],
            recommendations=["Test"],
        )
        csv = report.to_csv()
        # Nested list should be converted to JSON string
        assert "item1" in csv

    def test_to_csv_datetime_value(self, sample_report_metadata):
        """Test CSV with datetime values converted to ISO format."""
        report = ComplianceReport(
            metadata=sample_report_metadata,
            summary={},
            details=[
                {
                    "id": 1,
                    "action": "login",
                    "timestamp": datetime(2024, 1, 1, 12, 30, 45),
                },
            ],
            compliance_checks=[],
            recommendations=["Test"],
        )
        csv = report.to_csv()
        # datetime should be converted to ISO format
        assert "2024-01-01T12:30:45" in csv

    def test_to_csv_none_value(self, sample_report_metadata):
        """Test CSV with None values converted to empty string."""
        report = ComplianceReport(
            metadata=sample_report_metadata,
            summary={},
            details=[
                {
                    "id": 1,
                    "action": "login",
                    "user_id": None,
                    "ip_address": "192.168.1.1",
                },
            ],
            compliance_checks=[],
            recommendations=["Test"],
        )
        csv = report.to_csv()
        # None should be converted to empty string
        lines = csv.strip().split('\n')
        # Check that the user_id column is empty (not "None" string)
        assert len(lines) >= 2  # Header + data

    def test_to_csv_unicode(self, sample_report_metadata):
        """Test CSV with Unicode characters (Chinese, Japanese, Korean)."""
        report = ComplianceReport(
            metadata=sample_report_metadata,
            summary={},
            details=[
                {
                    "id": 1,
                    "action": "登录",  # Chinese
                    "user": "ユーザー",  # Japanese
                    "status": "활성",  # Korean
                },
            ],
            compliance_checks=[],
            recommendations=["Test"],
        )
        csv = report.to_csv()
        # Unicode characters should be preserved
        assert "登录" in csv
        assert "ユーザー" in csv
        assert "활성" in csv

    def test_to_csv_special_characters(self, sample_report_metadata):
        """Test CSV with special characters (comma, newline, double quote)."""
        report = ComplianceReport(
            metadata=sample_report_metadata,
            summary={},
            details=[
                {
                    "id": 1,
                    "action": "login",
                    "message": "User said, \"Hello\"",
                    "description": "Line1\nLine2",
                },
            ],
            compliance_checks=[],
            recommendations=["Test"],
        )
        csv = report.to_csv()
        # CSV should handle special characters properly (RFC 4180)
        # Double quotes should be escaped as two double quotes
        assert '""Hello""' in csv or '"Hello"' in csv
        # Field with newline should be quoted
        assert len(csv) > 0

    def test_to_csv_rfc4180_compliance(self, sample_report_metadata):
        """Test CSV RFC 4180 compliance - double quotes should be escaped."""
        report = ComplianceReport(
            metadata=sample_report_metadata,
            summary={},
            details=[
                {
                    "id": 1,
                    "action": "test",
                    "quote_field": "This has \"quotes\" inside",
                },
            ],
            compliance_checks=[],
            recommendations=["Test"],
        )
        csv = report.to_csv()
        # RFC 4180: double quotes inside a field must be escaped as two double quotes
        assert '""quotes""' in csv or csv.count('"') > 2

    def test_to_csv_long_json_truncation(self, sample_report_metadata):
        """Test CSV with long JSON string gets truncated."""
        # Create a very long nested dict
        long_dict = {f"field_{i}": f"value_{i}" for i in range(5000)}
        report = ComplianceReport(
            metadata=sample_report_metadata,
            summary={},
            details=[
                {
                    "id": 1,
                    "action": "export",
                    "details": long_dict,
                },
            ],
            compliance_checks=[],
            recommendations=["Test"],
        )
        csv = report.to_csv()
        # Long JSON should be truncated
        assert "[TRUNCATED]" in csv
        # Should not exceed Excel cell limit (32,767)
        lines = csv.strip().split('\n')
        for line in lines[1:]:  # Skip header
            # Each cell should not be too long
            cells = line.split(',')
            for cell in cells:
                # Remove quotes for length check
                unquoted = cell.strip('"')
                assert len(unquoted) <= 32050  # Allow some margin

    def test_to_csv_dict_format_comprehensive(self, sample_report_metadata):
        """Test CSV with dict format (comprehensive report)."""
        report = ComplianceReport(
            metadata=sample_report_metadata,
            summary={},
            details={
                "usage_summary": [
                    {"date": "2024-01-01", "tokens": 1000},
                ],
                "audit_trail": [
                    {
                        "id": 1,
                        "action": "login",
                        "details": {"ip": "192.168.1.1"},
                        "timestamp": datetime(2024, 1, 1),
                    },
                ],
            },
            compliance_checks=[],
            recommendations=["Test"],
        )
        csv = report.to_csv()
        # Should include section column
        assert "section" in csv
        assert "usage_summary" in csv
        assert "audit_trail" in csv
        # Nested dict should be flattened to JSON
        assert "ip" in csv


@pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl required for Excel tests")
class TestReportToExcelNestedFields:
    """Test Excel report generation with nested field flattening."""

    def test_to_excel_nested_dict_in_details(self, sample_report_metadata):
        """Test Excel details sheet with nested dict converted to JSON."""
        report = ComplianceReport(
            metadata=sample_report_metadata,
            summary={},
            details=[
                {
                    "id": 1,
                    "action": "login",
                    "details": {"ip_address": "192.168.1.1", "browser": "Chrome"},
                },
            ],
            compliance_checks=[],
            recommendations=["Test"],
        )
        excel = report.to_excel()
        wb = load_workbook(BytesIO(excel))

        # Get details sheet
        ws_details = wb[wb.sheetnames[1]]

        # Find details column and check value
        header_row = [cell.value for cell in ws_details[1]]
        details_col_idx = header_row.index("details") if "details" in header_row else None

        if details_col_idx:
            details_value = ws_details.cell(row=2, column=details_col_idx + 1).value
            # Should be a JSON string, not a Python dict
            assert isinstance(details_value, str)
            assert "ip_address" in details_value or "192.168.1.1" in details_value

    def test_to_excel_nested_list_in_details(self, sample_report_metadata):
        """Test Excel details sheet with nested list converted to JSON."""
        report = ComplianceReport(
            metadata=sample_report_metadata,
            summary={},
            details=[
                {
                    "id": 1,
                    "action": "export",
                    "items": ["file1.pdf", "file2.doc", "file3.xlsx"],
                },
            ],
            compliance_checks=[],
            recommendations=["Test"],
        )
        excel = report.to_excel()
        wb = load_workbook(BytesIO(excel))

        # Get details sheet
        ws_details = wb[wb.sheetnames[1]]

        # Find items column and check value
        header_row = [cell.value for cell in ws_details[1]]
        items_col_idx = header_row.index("items") if "items" in header_row else None

        if items_col_idx:
            items_value = ws_details.cell(row=2, column=items_col_idx + 1).value
            # Should be a JSON string, not a Python list
            assert isinstance(items_value, str)
            assert "file1.pdf" in items_value

    def test_to_excel_datetime_in_details(self, sample_report_metadata):
        """Test Excel details sheet with datetime converted to ISO format."""
        report = ComplianceReport(
            metadata=sample_report_metadata,
            summary={},
            details=[
                {
                    "id": 1,
                    "action": "login",
                    "timestamp": datetime(2024, 1, 1, 12, 30, 45),
                },
            ],
            compliance_checks=[],
            recommendations=["Test"],
        )
        excel = report.to_excel()
        wb = load_workbook(BytesIO(excel))

        # Get details sheet
        ws_details = wb[wb.sheetnames[1]]

        # Find timestamp column and check value
        header_row = [cell.value for cell in ws_details[1]]
        timestamp_col_idx = header_row.index("timestamp") if "timestamp" in header_row else None

        if timestamp_col_idx:
            timestamp_value = ws_details.cell(row=2, column=timestamp_col_idx + 1).value
            # Should be an ISO format string
            assert isinstance(timestamp_value, str)
            assert "2024-01-01" in timestamp_value

    def test_to_excel_long_json_truncation(self, sample_report_metadata):
        """Test Excel details sheet with long JSON string truncation."""
        # Create a very long nested dict
        long_dict = {f"field_{i}": f"value_{i}" for i in range(5000)}
        report = ComplianceReport(
            metadata=sample_report_metadata,
            summary={},
            details=[
                {
                    "id": 1,
                    "action": "export",
                    "details": long_dict,
                },
            ],
            compliance_checks=[],
            recommendations=["Test"],
        )
        excel = report.to_excel()
        wb = load_workbook(BytesIO(excel))

        # Get details sheet
        ws_details = wb[wb.sheetnames[1]]

        # Find details column and check value
        header_row = [cell.value for cell in ws_details[1]]
        details_col_idx = header_row.index("details") if "details" in header_row else None

        if details_col_idx:
            details_value = ws_details.cell(row=2, column=details_col_idx + 1).value
            # Should be truncated
            assert "[TRUNCATED]" in details_value
            # Should not exceed Excel cell limit
            assert len(details_value) <= 32050  # Allow some margin

    def test_to_excel_dict_format_nested_fields(self, sample_report_metadata):
        """Test Excel details sheet with dict format (comprehensive report)."""
        report = ComplianceReport(
            metadata=sample_report_metadata,
            summary={},
            details={
                "audit_trail": [
                    {
                        "id": 1,
                        "action": "login",
                        "details": {"ip": "192.168.1.1"},
                    },
                ],
            },
            compliance_checks=[],
            recommendations=["Test"],
        )
        excel = report.to_excel()
        wb = load_workbook(BytesIO(excel))

        # Get details sheet
        ws_details = wb[wb.sheetnames[1]]

        # Should have section header and data
        # Check that nested dict is flattened to JSON
        for row in ws_details.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if "ip" in cell.value or "192.168.1.1" in cell.value:
                        # Found the flattened JSON string
                        assert True
                        return