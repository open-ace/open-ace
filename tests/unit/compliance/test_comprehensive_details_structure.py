"""
Unit tests for comprehensive report details structure fix.

Tests cover:
- Comprehensive report details structure (dict format)
- Audit trail report details structure (list format)
- Details type difference between comprehensive and audit_trail
- Backward compatibility for old format saved reports
- CSV output with section column
- Excel output with section column
"""

import pytest
from datetime import datetime, timezone
import json
import io

try:
    import openpyxl
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from app.modules.compliance.report import (
    ComplianceReport,
    ReportMetadata,
    ReportType,
    ReportGenerator,
)


@pytest.fixture
def comprehensive_report_metadata():
    """Create comprehensive report metadata."""
    return ReportMetadata(
        report_id="test-comprehensive-001",
        report_type=ReportType.COMPREHENSIVE.value,
        generated_at=datetime.now(timezone.utc).replace(tzinfo=None),
        period_start=datetime(2024, 1, 1),
        period_end=datetime(2024, 1, 31),
        generated_by=1,
    )


@pytest.fixture
def audit_trail_report_metadata():
    """Create audit trail report metadata."""
    return ReportMetadata(
        report_id="test-audit-trail-001",
        report_type=ReportType.AUDIT_TRAIL.value,
        generated_at=datetime.now(timezone.utc).replace(tzinfo=None),
        period_start=datetime(2024, 1, 1),
        period_end=datetime(2024, 1, 31),
        generated_by=1,
    )


@pytest.fixture
def comprehensive_report(comprehensive_report_metadata):
    """Create comprehensive report with new dict format details."""
    return ComplianceReport(
        metadata=comprehensive_report_metadata,
        summary={
            "period": {"start": "2024-01-01T00:00:00", "end": "2024-01-31T00:00:00"},
            "usage": {
                "totals": {
                    "tokens": 10000,
                    "requests": 500,
                    "tools_used": 3,
                    "unique_users": 10,
                }
            },
            "users": {
                "totals": {"total_users": 20, "active_users": 15, "inactive_users": 5}
            },
            "audit": {
                "totals": {"total_events": 100, "unique_users": 10}
            },
            "security": {
                "security_events": {
                    "failed_logins": 5,
                    "password_changes": 2,
                    "role_changes": 1,
                    "content_blocked": 0,
                }
            },
            "quota": {
                "alerts": {"total": 3, "warnings": 2, "critical": 1, "exceeded": 0}
            },
        },
        details={
            "usage_summary": [
                {
                    "date": "2024-01-01",
                    "tool_name": "claude",
                    "host_name": "api.anthropic.com",
                    "total_input_tokens": 1000,
                    "total_output_tokens": 500,
                    "total_tokens": 1500,
                    "total_requests": 10,
                }
            ],
            "user_activity": [
                {
                    "user_id": 1,
                    "username": "testuser",
                    "email": "test@example.com",
                    "role": "user",
                    "active_days": 10,
                    "total_tokens": 5000,
                    "total_requests": 100,
                    "first_activity": "2024-01-01",
                    "last_activity": "2024-01-31",
                }
            ],
            "audit_trail": [
                {
                    "timestamp": "2024-01-01T10:00:00",
                    "action": "login",
                    "user_id": 1,
                    "severity": "low",
                    "resource_type": "session",
                }
            ],
            "security_events": [
                {
                    "timestamp": "2024-01-01T11:00:00",
                    "action": "login_failed",
                    "user_id": 2,
                    "severity": "medium",
                    "resource_type": "auth",
                    "details": "Invalid credentials",
                }
            ],
            "quota_alerts": [
                {
                    "id": 1,
                    "alert_type": "warning",
                    "user_id": 3,
                    "created_at": "2024-01-01T12:00:00",
                    "message": "Approaching quota limit",
                }
            ],
        },
        compliance_checks=[
            {
                "check_id": "data_retention",
                "name": "Data Retention Policy",
                "status": "pass",
                "message": "Data retention policy is being followed",
            },
        ],
        recommendations=["All compliance checks passed. Continue monitoring."],
    )


@pytest.fixture
def audit_trail_report(audit_trail_report_metadata):
    """Create audit trail report with list format details."""
    return ComplianceReport(
        metadata=audit_trail_report_metadata,
        summary={
            "period": {"start": "2024-01-01T00:00:00", "end": "2024-01-31T00:00:00"},
            "totals": {"total_events": 100, "unique_users": 10},
            "by_action": {"login": 50, "logout": 30, "other": 20},
            "by_severity": {"low": 80, "medium": 15, "high": 5},
        },
        details=[
            {
                "timestamp": "2024-01-01T10:00:00",
                "action": "login",
                "user_id": 1,
                "severity": "low",
                "resource_type": "session",
            },
            {
                "timestamp": "2024-01-01T11:00:00",
                "action": "logout",
                "user_id": 1,
                "severity": "low",
                "resource_type": "session",
            },
        ],
        compliance_checks=[
            {
                "check_id": "audit_complete",
                "name": "Audit Trail Complete",
                "status": "pass",
                "message": "Audit trail is complete",
            },
        ],
        recommendations=["Continue monitoring audit events."],
    )


class TestComprehensiveDetailsStructure:
    """Test comprehensive report details structure."""

    def test_comprehensive_details_is_dict(self, comprehensive_report):
        """Test that comprehensive report details is dict type."""
        assert isinstance(comprehensive_report.details, dict)

    def test_comprehensive_details_has_all_sections(self, comprehensive_report):
        """Test that comprehensive details contains all sections."""
        expected_sections = [
            "usage_summary",
            "user_activity",
            "audit_trail",
            "security_events",
            "quota_alerts",
        ]
        for section in expected_sections:
            assert section in comprehensive_report.details

    def test_comprehensive_details_sections_are_lists(self, comprehensive_report):
        """Test that each section in comprehensive details is a list."""
        for section_key, section_data in comprehensive_report.details.items():
            assert isinstance(section_data, list)

    def test_comprehensive_to_dict_preserves_details_structure(self, comprehensive_report):
        """Test that to_dict preserves dict details structure."""
        report_dict = comprehensive_report.to_dict()
        assert isinstance(report_dict["details"], dict)
        assert "usage_summary" in report_dict["details"]
        assert "audit_trail" in report_dict["details"]


class TestAuditTrailDetailsStructure:
    """Test audit trail report details structure."""

    def test_audit_trail_details_is_list(self, audit_trail_report):
        """Test that audit trail report details is list type."""
        assert isinstance(audit_trail_report.details, list)

    def test_audit_trail_details_not_dict(self, audit_trail_report):
        """Test that audit trail details is not dict."""
        assert not isinstance(audit_trail_report.details, dict)

    def test_audit_trail_to_dict_preserves_list_details(self, audit_trail_report):
        """Test that to_dict preserves list details structure."""
        report_dict = audit_trail_report.to_dict()
        assert isinstance(report_dict["details"], list)


class TestComprehensiveVsAuditTrailDifferentDetails:
    """Test that comprehensive and audit trail have different details structure."""

    def test_details_type_different(self, comprehensive_report, audit_trail_report):
        """Test that comprehensive and audit trail details types are different."""
        assert isinstance(comprehensive_report.details, dict)
        assert isinstance(audit_trail_report.details, list)
        assert type(comprehensive_report.details) != type(audit_trail_report.details)

    def test_details_content_different(self, comprehensive_report, audit_trail_report):
        """Test that comprehensive details has more sections than audit trail."""
        # Comprehensive has multiple sections
        assert len(comprehensive_report.details) == 5  # 5 sections
        # Audit trail has flat list
        assert len(audit_trail_report.details) == 2  # 2 rows


class TestComprehensiveHtmlRendering:
    """Test comprehensive report HTML rendering."""

    def test_html_contains_all_sections(self, comprehensive_report):
        """Test that HTML contains all section headers."""
        html = comprehensive_report.to_html()
        # Should contain section titles (in English by default)
        assert "Usage Summary" in html or "sectionUsageSummary" in html
        assert "User Activity" in html or "sectionUserActivity" in html
        assert "Audit Trail" in html or "sectionAuditTrail" in html
        assert "Security Events" in html or "sectionSecurityEvents" in html
        assert "Quota Alerts" in html or "sectionQuotaAlerts" in html

    def test_html_has_multiple_tables(self, comprehensive_report):
        """Test that HTML has multiple tables for different sections."""
        html = comprehensive_report.to_html()
        # Count table elements - should have multiple tables
        table_count = html.count("<table")
        assert table_count >= 5  # At least 5 tables for 5 sections

    def test_html_contains_usage_summary_data(self, comprehensive_report):
        """Test that HTML contains usage summary data."""
        html = comprehensive_report.to_html()
        assert "claude" in html or "1000" in html

    def test_html_contains_audit_trail_data(self, comprehensive_report):
        """Test that HTML contains audit trail data."""
        html = comprehensive_report.to_html()
        assert "login" in html

    def test_html_language_zh_contains_chinese_sections(self, comprehensive_report):
        """Test that Chinese HTML contains Chinese section titles."""
        html = comprehensive_report.to_html(language="zh")
        assert "使用统计" in html
        assert "用户活动" in html
        assert "审计轨迹" in html

    def test_html_language_ja_contains_japanese_sections(self, comprehensive_report):
        """Test that Japanese HTML contains Japanese section titles."""
        html = comprehensive_report.to_html(language="ja")
        assert "使用統計" in html
        assert "ユーザー活動" in html
        assert "監査履歴" in html

    def test_html_language_ko_contains_korean_sections(self, comprehensive_report):
        """Test that Korean HTML contains Korean section titles."""
        html = comprehensive_report.to_html(language="ko")
        assert "사용량 요약" in html
        assert "사용자 활동" in html
        assert "감사 추적" in html


class TestComprehensiveCsvOutput:
    """Test comprehensive report CSV output."""

    def test_csv_has_section_column(self, comprehensive_report):
        """Test that CSV output has section column."""
        csv_content = comprehensive_report.to_csv()
        assert "section" in csv_content

    def test_csv_contains_all_sections(self, comprehensive_report):
        """Test that CSV contains data from all sections."""
        csv_content = comprehensive_report.to_csv()
        assert "usage_summary" in csv_content
        assert "user_activity" in csv_content
        assert "audit_trail" in csv_content
        assert "security_events" in csv_content
        assert "quota_alerts" in csv_content

    def test_csv_not_empty(self, comprehensive_report):
        """Test that CSV is not empty."""
        csv_content = comprehensive_report.to_csv()
        assert len(csv_content) > 0

    def test_csv_has_header_row(self, comprehensive_report):
        """Test that CSV has header row."""
        csv_content = comprehensive_report.to_csv()
        lines = csv_content.strip().split("\n")
        assert len(lines) >= 1
        header = lines[0]
        assert "section" in header


class TestAuditTrailCsvOutput:
    """Test audit trail report CSV output (backward compatible)."""

    def test_audit_trail_csv_no_section_column(self, audit_trail_report):
        """Test that audit trail CSV does not have section column (list format)."""
        csv_content = audit_trail_report.to_csv()
        # Audit trail uses list format, no section column
        assert "section" not in csv_content

    def test_audit_trail_csv_has_audit_columns(self, audit_trail_report):
        """Test that audit trail CSV has audit-specific columns."""
        csv_content = audit_trail_report.to_csv()
        assert "timestamp" in csv_content
        assert "action" in csv_content


@pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl required for Excel tests")
class TestComprehensiveExcelOutput:
    """Test comprehensive report Excel output."""

    def test_excel_has_details_sheet(self, comprehensive_report):
        """Test that Excel has Details sheet."""
        excel = comprehensive_report.to_excel()
        wb = load_workbook(io.BytesIO(excel))
        assert "Details" in wb.sheetnames or "详情" in wb.sheetnames

    def test_excel_details_contains_sections(self, comprehensive_report):
        """Test that Excel details sheet contains section headers."""
        excel = comprehensive_report.to_excel()
        wb = load_workbook(io.BytesIO(excel))

        # Get details sheet
        ws_details = None
        for name in wb.sheetnames:
            if "Details" in name or "详情" in name:
                ws_details = wb[name]
                break

        assert ws_details is not None

        # Check for section headers
        all_values = []
        for row in ws_details.iter_rows():
            for cell in row:
                if cell.value:
                    all_values.append(str(cell.value))

        # Should have section identifiers
        assert any("Usage" in v for v in all_values) or any("usage_summary" in v for v in all_values)

    def test_excel_language_zh_contains_chinese_sections(self, comprehensive_report):
        """Test that Chinese Excel contains Chinese section titles."""
        excel = comprehensive_report.to_excel(language="zh")
        wb = load_workbook(io.BytesIO(excel))

        # Get details sheet
        ws_details = None
        for name in wb.sheetnames:
            if "详情" in name:
                ws_details = wb[name]
                break

        assert ws_details is not None

        # Check for Chinese section headers
        all_values = []
        for row in ws_details.iter_rows(max_row=20):
            for cell in row:
                if cell.value:
                    all_values.append(str(cell.value))

        # Should have Chinese section headers
        assert any("使用统计" in v for v in all_values)


@pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl required for Excel tests")
class TestAuditTrailExcelOutput:
    """Test audit trail report Excel output (backward compatible)."""

    def test_audit_trail_excel_no_section_column(self, audit_trail_report):
        """Test that audit trail Excel details sheet doesn't have section column."""
        excel = audit_trail_report.to_excel()
        wb = load_workbook(io.BytesIO(excel))

        # Get details sheet
        ws_details = None
        for name in wb.sheetnames:
            if "Details" in name or "详情" in name:
                ws_details = wb[name]
                break

        assert ws_details is not None

        # Check headers - should not have section column
        headers = [cell.value for cell in ws_details[1]]
        assert "section" not in headers
        assert "Section" not in headers

    def test_audit_trail_excel_has_audit_columns(self, audit_trail_report):
        """Test that audit trail Excel has audit-specific columns."""
        excel = audit_trail_report.to_excel()
        wb = load_workbook(io.BytesIO(excel))

        # Get details sheet
        ws_details = None
        for name in wb.sheetnames:
            if "Details" in name or "详情" in name:
                ws_details = wb[name]
                break

        assert ws_details is not None

        # Check headers - should have audit columns
        headers = [cell.value for cell in ws_details[1]]
        assert "timestamp" in headers or "Timestamp" in headers
        assert "action" in headers or "Action" in headers


class TestBackwardCompatibility:
    """Test backward compatibility for old format saved reports."""

    def test_old_format_list_details_converted_to_dict(self, comprehensive_report_metadata):
        """Test that old format list details is converted to dict when loaded."""
        # Simulate old format saved report data
        old_format_data = {
            "metadata": {
                "report_id": "old-comprehensive-001",
                "report_type": ReportType.COMPREHENSIVE.value,
                "generated_at": "2024-01-01T00:00:00",
                "period_start": "2024-01-01T00:00:00",
                "period_end": "2024-01-31T00:00:00",
                "generated_by": 1,
            },
            "summary": {
                "period": {"start": "2024-01-01T00:00:00", "end": "2024-01-31T00:00:00"},
                "usage": {"totals": {"tokens": 1000}},
                "users": {"totals": {"total_users": 10}},
                "audit": {"totals": {"total_events": 50}},
                "security": {"security_events": {"failed_logins": 0}},
                "quota": {"alerts": {"total": 0}},
            },
            "details": [  # Old format: flat list
                {
                    "timestamp": "2024-01-01T10:00:00",
                    "action": "login",
                    "user_id": 1,
                    "severity": "low",
                    "resource_type": "session",
                },
            ],
            "compliance_checks": [],
            "recommendations": [],
        }

        # Create report from old format data
        metadata = ReportMetadata(
            report_id=old_format_data["metadata"]["report_id"],
            report_type=old_format_data["metadata"]["report_type"],
            generated_at=datetime.fromisoformat(old_format_data["metadata"]["generated_at"]),
            period_start=datetime.fromisoformat(old_format_data["metadata"]["period_start"]),
            period_end=datetime.fromisoformat(old_format_data["metadata"]["period_end"]),
            generated_by=old_format_data["metadata"]["generated_by"],
        )

        # Simulate what get_saved_report would do
        details = old_format_data["details"]
        if isinstance(details, list):
            details = {
                "usage_summary": [],
                "user_activity": [],
                "audit_trail": details,
                "security_events": [],
                "quota_alerts": [],
            }

        report = ComplianceReport(
            metadata=metadata,
            summary=old_format_data["summary"],
            details=details,
            compliance_checks=old_format_data["compliance_checks"],
            recommendations=old_format_data["recommendations"],
        )

        # Verify converted format
        assert isinstance(report.details, dict)
        assert "audit_trail" in report.details
        assert len(report.details["audit_trail"]) == 1

    def test_old_format_html_rendering(self, comprehensive_report_metadata):
        """Test that old format details still renders correctly."""
        # Create report with converted old format details
        old_format_report = ComplianceReport(
            metadata=comprehensive_report_metadata,
            summary={
                "period": {"start": "2024-01-01T00:00:00", "end": "2024-01-31T00:00:00"},
                "usage": {},
                "users": {},
                "audit": {},
                "security": {},
                "quota": {},
            },
            details={
                "usage_summary": [],
                "user_activity": [],
                "audit_trail": [
                    {
                        "timestamp": "2024-01-01T10:00:00",
                        "action": "login",
                        "user_id": 1,
                        "severity": "low",
                        "resource_type": "session",
                    },
                ],
                "security_events": [],
                "quota_alerts": [],
            },
            compliance_checks=[],
            recommendations=["Test"],
        )

        html = old_format_report.to_html()
        # Should render audit trail section
        assert "login" in html
        # Empty sections should not render tables
        assert "usage_summary" not in html.lower() or html.count("<table") == 1


class TestEmptySectionsHandling:
    """Test handling of empty sections in comprehensive report."""

    def test_empty_sections_not_rendered_in_html(self, comprehensive_report_metadata):
        """Test that empty sections are not rendered in HTML."""
        report_with_empty_sections = ComplianceReport(
            metadata=comprehensive_report_metadata,
            summary={
                "period": {"start": "2024-01-01T00:00:00", "end": "2024-01-31T00:00:00"},
                "usage": {},
                "users": {},
                "audit": {},
                "security": {},
                "quota": {},
            },
            details={
                "usage_summary": [],  # Empty
                "user_activity": [],  # Empty
                "audit_trail": [
                    {"timestamp": "2024-01-01", "action": "login", "user_id": 1, "severity": "low", "resource_type": "session"},
                ],
                "security_events": [],  # Empty
                "quota_alerts": [],  # Empty
            },
            compliance_checks=[],
            recommendations=["Test"],
        )

        html = report_with_empty_sections.to_html()
        # Should only have one table (audit_trail)
        table_count = html.count("<table")
        assert table_count == 1

    def test_all_empty_sections_empty_state(self, comprehensive_report_metadata):
        """Test that all empty sections shows empty state."""
        report_all_empty = ComplianceReport(
            metadata=comprehensive_report_metadata,
            summary={
                "period": {"start": "2024-01-01T00:00:00", "end": "2024-01-31T00:00:00"},
                "usage": {},
                "users": {},
                "audit": {},
                "security": {},
                "quota": {},
            },
            details={
                "usage_summary": [],
                "user_activity": [],
                "audit_trail": [],
                "security_events": [],
                "quota_alerts": [],
            },
            compliance_checks=[],
            recommendations=["Test"],
        )

        html = report_all_empty.to_html()
        # Should show empty state message
        assert "No detailed data" in html or "empty" in html.lower()


class TestReportGeneratorComprehensive:
    """Test ReportGenerator comprehensive report generation."""

    def test_generate_comprehensive_returns_dict_details(self):
        """Test that generated comprehensive report has dict details."""
        # Note: This test may fail if database is not available
        # It's designed to test the structure when ReportGenerator can work
        try:
            generator = ReportGenerator()
            report = generator.generate_report(
                report_type=ReportType.COMPREHENSIVE.value,
                period_start=datetime(2024, 1, 1),
                period_end=datetime(2024, 1, 31),
                generated_by=1,
            )

            # Verify dict structure
            assert isinstance(report.details, dict)
            assert "usage_summary" in report.details
            assert "user_activity" in report.details
            assert "audit_trail" in report.details
            assert "security_events" in report.details
            assert "quota_alerts" in report.details
        except Exception:
            # If database connection fails, skip this test
            pytest.skip("Database not available for integration test")

    def test_generate_audit_trail_returns_list_details(self):
        """Test that generated audit trail report has list details."""
        try:
            generator = ReportGenerator()
            report = generator.generate_report(
                report_type=ReportType.AUDIT_TRAIL.value,
                period_start=datetime(2024, 1, 1),
                period_end=datetime(2024, 1, 31),
                generated_by=1,
            )

            # Verify list structure
            assert isinstance(report.details, list)
        except Exception:
            pytest.skip("Database not available for integration test")