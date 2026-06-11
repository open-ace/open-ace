"""
Unit tests for ComplianceReport to_html() and to_excel() methods.

Tests cover:
- HTML template rendering with different report types
- Excel generation with multiple worksheets
- XSS protection through Jinja2 autoescape
- Empty report handling
- Multi-language support
"""

import pytest
from datetime import datetime, timezone

from app.modules.compliance.report import (
    ComplianceReport,
    ReportMetadata,
    ReportType,
)


@pytest.fixture
def sample_report_metadata():
    """Create sample report metadata."""
    return ReportMetadata(
        report_id="test-report-001",
        report_type=ReportType.USAGE_SUMMARY.value,
        generated_at=datetime.now(timezone.utc).replace(tzinfo=None),
        period_start=datetime(2024, 1, 1),
        period_end=datetime(2024, 1, 31),
        generated_by=1,
    )


@pytest.fixture
def sample_report(sample_report_metadata):
    """Create sample compliance report."""
    return ComplianceReport(
        metadata=sample_report_metadata,
        summary={
            "period": {"start": "2024-01-01", "end": "2024-01-31", "days": 31},
            "totals": {
                "tokens": 10000,
                "requests": 500,
                "tools_used": 3,
                "unique_users": 10,
            },
        },
        details=[
            {
                "date": "2024-01-01",
                "tool_name": "claude",
                "host_name": "api.anthropic.com",
                "total_input_tokens": 1000,
                "total_output_tokens": 500,
                "total_tokens": 1500,
                "total_requests": 10,
            },
            {
                "date": "2024-01-02",
                "tool_name": "gpt",
                "host_name": "api.openai.com",
                "total_input_tokens": 2000,
                "total_output_tokens": 1000,
                "total_tokens": 3000,
                "total_requests": 20,
            },
        ],
        compliance_checks=[
            {
                "check_id": "data_retention",
                "name": "Data Retention Policy",
                "status": "pass",
                "message": "Data retention policy is being followed",
            },
            {
                "check_id": "inactive_users",
                "name": "Inactive User Review",
                "status": "warning",
                "message": "5 inactive users found. Review access.",
            },
        ],
        recommendations=[
            "All compliance checks passed. Continue monitoring.",
            "Review: Inactive User Review - 5 inactive users found.",
        ],
    )


class TestReportToHtml:
    """Test HTML report generation."""

    def test_to_html_returns_string(self, sample_report):
        """Test that to_html returns a string."""
        html = sample_report.to_html()
        assert isinstance(html, str)
        assert len(html) > 0

    def test_to_html_contains_metadata(self, sample_report):
        """Test that HTML contains report metadata."""
        html = sample_report.to_html()
        assert "test-report-001" in html
        assert "2024-01-01" in html
        assert "2024-01-31" in html

    def test_to_html_contains_summary(self, sample_report):
        """Test that HTML contains summary data."""
        html = sample_report.to_html()
        assert "10000" in html or "tokens" in html.lower()
        assert "500" in html or "requests" in html.lower()

    def test_to_html_contains_details_table(self, sample_report):
        """Test that HTML contains details table."""
        html = sample_report.to_html()
        assert "<table" in html
        assert "<thead" in html
        assert "<tbody" in html

    def test_to_html_contains_compliance_checks(self, sample_report):
        """Test that HTML contains compliance checks."""
        html = sample_report.to_html()
        assert "Data Retention Policy" in html
        assert "Inactive User Review" in html
        # Check for status icons/badges
        assert "pass" in html.lower() or "Pass" in html
        assert "warning" in html.lower() or "Warning" in html

    def test_to_html_contains_recommendations(self, sample_report):
        """Test that HTML contains recommendations."""
        html = sample_report.to_html()
        assert "recommendations" in html.lower() or "Recommendations" in html

    def test_to_html_xss_protection(self, sample_report_metadata):
        """Test that XSS attempts are escaped."""
        malicious_report = ComplianceReport(
            metadata=sample_report_metadata,
            summary={"test": "<script>alert('xss')</script>"},
            details=[{"name": "<img src=x onerror=alert('xss')>"}],
            compliance_checks=[],
            recommendations=["<a href='javascript:alert(1)'>click</a>"],
        )
        html = malicious_report.to_html()
        # XSS payloads should be escaped
        assert "<script>" not in html
        assert "onerror=" not in html
        assert "javascript:" not in html
        # Escaped versions should be present
        assert "&lt;script" in html or "alert" in html

    def test_to_html_empty_details(self, sample_report_metadata):
        """Test HTML with empty details."""
        empty_report = ComplianceReport(
            metadata=sample_report_metadata,
            summary={"totals": {"tokens": 0, "requests": 0}},
            details=[],
            compliance_checks=[],
            recommendations=["No data available"],
        )
        html = empty_report.to_html()
        assert "No usage data" in html or "No data" in html or "empty" in html.lower()

    def test_to_html_language_en(self, sample_report):
        """Test English language HTML."""
        html = sample_report.to_html(language="en")
        assert "Compliance Report" in html or "Usage Summary" in html

    def test_to_html_language_zh(self, sample_report):
        """Test Chinese language HTML."""
        html = sample_report.to_html(language="zh")
        # Chinese translation for report title
        assert "使用汇总报告" in html or "合规报告" in html

    def test_to_html_language_ja(self, sample_report):
        """Test Japanese language HTML."""
        html = sample_report.to_html(language="ja")
        # Japanese translation for report title
        assert "使用量サマリー" in html or "コンプライアンスレポート" in html

    def test_to_html_language_ko(self, sample_report):
        """Test Korean language HTML."""
        html = sample_report.to_html(language="ko")
        # Korean translation for report title
        assert "사용량 요약 보고서" in html or "준수 보고서" in html

    def test_to_html_contains_print_styles(self, sample_report):
        """Test that HTML contains print optimization styles."""
        html = sample_report.to_html()
        assert "@media print" in html
        assert "page-break-inside" in html

    def test_to_html_contains_watermark(self, sample_report):
        """Test that HTML contains watermark."""
        html = sample_report.to_html()
        assert "Generated by" in html
        assert "Open ACE" in html

    def test_to_html_different_report_types(self):
        """Test HTML for different report types."""
        report_types = [
            ReportType.USER_ACTIVITY.value,
            ReportType.AUDIT_TRAIL.value,
            ReportType.SECURITY.value,
            ReportType.COMPREHENSIVE.value,
        ]

        for report_type in report_types:
            metadata = ReportMetadata(
                report_id=f"test-{report_type}",
                report_type=report_type,
                generated_at=datetime.now(timezone.utc).replace(tzinfo=None),
                period_start=datetime(2024, 1, 1),
                period_end=datetime(2024, 1, 31),
            )
            report = ComplianceReport(
                metadata=metadata,
                summary={},
                details=[],
                compliance_checks=[],
                recommendations=["Test recommendation"],
            )
            html = report.to_html()
            assert isinstance(html, str)
            assert len(html) > 0


class TestReportToExcel:
    """Test Excel report generation."""

    def test_to_excel_returns_bytes(self, sample_report):
        """Test that to_excel returns bytes."""
        excel = sample_report.to_excel()
        assert isinstance(excel, bytes)
        assert len(excel) > 0

    def test_to_excel_valid_format(self, sample_report):
        """Test that Excel file is valid xlsx format."""
        import io
        from openpyxl import load_workbook

        excel = sample_report.to_excel()
        wb = load_workbook(io.BytesIO(excel))

        # Should have multiple sheets
        assert len(wb.sheetnames) >= 3
        assert "Summary" in wb.sheetnames or "汇总" in wb.sheetnames
        assert "Details" in wb.sheetnames or "详情" in wb.sheetnames

    def test_to_excel_contains_metadata(self, sample_report):
        """Test that Excel contains metadata."""
        import io
        from openpyxl import load_workbook

        excel = sample_report.to_excel()
        wb = load_workbook(io.BytesIO(excel))
        ws = wb.active

        # Check metadata cells
        assert "test-report-001" in str(ws["B1"].value)

    def test_to_excel_contains_summary(self, sample_report):
        """Test that Excel contains summary data."""
        import io
        from openpyxl import load_workbook

        excel = sample_report.to_excel()
        wb = load_workbook(io.BytesIO(excel))

        # Summary sheet
        ws_summary = wb.active
        # Should have summary section
        assert ws_summary.max_row > 5

    def test_to_excel_contains_details(self, sample_report):
        """Test that Excel contains details."""
        import io
        from openpyxl import load_workbook

        excel = sample_report.to_excel()
        wb = load_workbook(io.BytesIO(excel))

        # Details sheet
        ws_details = wb[wb.sheetnames[1]]  # Second sheet is details

        # Should have header and data rows
        assert ws_details.max_row >= 2  # Header + at least 1 data row

    def test_to_excel_contains_compliance_checks(self, sample_report):
        """Test that Excel contains compliance checks."""
        import io
        from openpyxl import load_workbook

        excel = sample_report.to_excel()
        wb = load_workbook(io.BytesIO(excel))

        # Find compliance checks sheet
        sheet_names = wb.sheetnames
        checks_sheet = None
        for name in sheet_names:
            if "Checks" in name or "Compliance" in name or "检查" in name:
                checks_sheet = wb[name]
                break

        assert checks_sheet is not None
        # Should have header and check rows
        assert checks_sheet.max_row >= 2

    def test_to_excel_empty_details(self, sample_report_metadata):
        """Test Excel with empty details."""
        import io
        from openpyxl import load_workbook

        empty_report = ComplianceReport(
            metadata=sample_report_metadata,
            summary={},
            details=[],
            compliance_checks=[],
            recommendations=["No data available"],
        )
        excel = empty_report.to_excel()
        wb = load_workbook(io.BytesIO(excel))

        # Details sheet should exist but have no data rows
        ws_details = wb[wb.sheetnames[1]]
        assert ws_details.max_row == 1  # Only header or "no data" message

    def test_to_excel_language_en(self, sample_report):
        """Test English language Excel."""
        import io
        from openpyxl import load_workbook

        excel = sample_report.to_excel(language="en")
        wb = load_workbook(io.BytesIO(excel))

        # Sheet names should be in English
        assert "Summary" in wb.sheetnames or any("Summary" in s for s in wb.sheetnames)

    def test_to_excel_language_zh(self, sample_report):
        """Test Chinese language Excel."""
        import io
        from openpyxl import load_workbook

        excel = sample_report.to_excel(language="zh")
        wb = load_workbook(io.BytesIO(excel))

        # Sheet names should be in Chinese
        assert "汇总" in wb.sheetnames or any("汇总" in s for s in wb.sheetnames)

    def test_to_excel_styling(self, sample_report):
        """Test Excel styling applied."""
        import io
        from openpyxl import load_workbook

        excel = sample_report.to_excel()
        wb = load_workbook(io.BytesIO(excel))

        ws_summary = wb.active
        # First row metadata should have bold font
        cell = ws_summary["A1"]
        assert cell.font.bold

    def test_to_excel_different_report_types(self):
        """Test Excel for different report types."""
        import io
        from openpyxl import load_workbook

        report_types = [
            ReportType.USER_ACTIVITY.value,
            ReportType.AUDIT_TRAIL.value,
            ReportType.SECURITY.value,
            ReportType.COMPREHENSIVE.value,
        ]

        for report_type in report_types:
            metadata = ReportMetadata(
                report_id=f"test-{report_type}",
                report_type=report_type,
                generated_at=datetime.now(timezone.utc).replace(tzinfo=None),
                period_start=datetime(2024, 1, 1),
                period_end=datetime(2024, 1, 31),
            )
            report = ComplianceReport(
                metadata=metadata,
                summary={},
                details=[],
                compliance_checks=[],
                recommendations=["Test"],
            )
            excel = report.to_excel()
            wb = load_workbook(io.BytesIO(excel))
            assert len(wb.sheetnames) >= 1


class TestReportTemplates:
    """Test HTML template existence and validity."""

    def test_templates_directory_exists(self):
        """Test that templates directory exists."""
        import os
        from app.modules.compliance.report import ComplianceReport

        template_dir = os.path.join(
            os.path.dirname(ComplianceReport.__module__.replace(".", "/")),
            "templates"
        )
        # Template directory should exist
        assert os.path.exists(template_dir)

    def test_base_template_exists(self):
        """Test that base.html template exists."""
        import os

        template_path = os.path.join(
            os.path.dirname(__file__), "..", "..", "..",
            "app", "modules", "compliance", "templates", "base.html"
        )
        assert os.path.exists(template_path)

    def test_component_templates_exist(self):
        """Test that component templates exist."""
        import os

        template_dir = os.path.join(
            os.path.dirname(__file__), "..", "..", "..",
            "app", "modules", "compliance", "templates", "components"
        )

        components = ["header.html", "summary_table.html", "compliance_checks.html", "recommendations.html"]
        for component in components:
            assert os.path.exists(os.path.join(template_dir, component))

    def test_report_templates_exist(self):
        """Test that report templates exist."""
        import os

        template_dir = os.path.join(
            os.path.dirname(__file__), "..", "..", "..",
            "app", "modules", "compliance", "templates", "reports"
        )

        reports = ["usage_summary.html", "user_activity.html", "audit_trail.html", "security.html", "comprehensive.html"]
        for report in reports:
            assert os.path.exists(os.path.join(template_dir, report))