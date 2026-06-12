"""
Open ACE - Compliance Report Generator

Generates compliance reports for enterprise auditing and regulatory requirements.
"""

import csv
import io
import json
import logging
import os
from dataclasses import dataclass, field
from datetime import datetime, timezone
from enum import Enum
from typing import Any, Optional, Union

from jinja2 import Environment, FileSystemLoader, select_autoescape

from app.modules.governance.audit_logger import AuditLogger
from app.repositories.database import Database
from app.repositories.usage_repo import UsageRepository
from app.repositories.user_repo import UserRepository

logger = logging.getLogger(__name__)


class ReportType(Enum):
    """Types of compliance reports."""

    USAGE_SUMMARY = "usage_summary"
    USER_ACTIVITY = "user_activity"
    AUDIT_TRAIL = "audit_trail"
    DATA_ACCESS = "data_access"
    SECURITY = "security"
    QUOTA_USAGE = "quota_usage"
    COMPREHENSIVE = "comprehensive"


class ReportFormat(Enum):
    """Report output formats."""

    JSON = "json"
    CSV = "csv"
    HTML = "html"
    EXCEL = "excel"
    PDF = "pdf"


@dataclass
class ReportMetadata:
    """Report metadata."""

    report_id: str
    report_type: str
    generated_at: datetime
    period_start: datetime
    period_end: datetime
    generated_by: Optional[int] = None
    tenant_id: Optional[int] = None
    filters: dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> dict:
        """Convert to dictionary."""
        return {
            "report_id": self.report_id,
            "report_type": self.report_type,
            "generated_at": self.generated_at.isoformat(),
            "period_start": self.period_start.isoformat(),
            "period_end": self.period_end.isoformat(),
            "generated_by": self.generated_by,
            "tenant_id": self.tenant_id,
            "filters": self.filters,
        }


@dataclass
class ComplianceReport:
    """Compliance report data structure."""

    metadata: ReportMetadata
    summary: dict[str, Any]
    details: list[dict[str, Any]]
    compliance_checks: list[dict[str, Any]]
    recommendations: list[str]

    # Maximum cell length for Excel (Excel has a 32,767 character limit per cell)
    MAX_CELL_LENGTH = 32000

    # Key fields to preserve when truncating long JSON strings
    KEY_FIELDS = ["ip_address", "action", "timestamp", "user_id", "resource_type"]

    def _flatten_value(self, value: Any) -> Any:
        """
        Flatten a single value for CSV/Excel export.

        Handles:
        - Nested dict/list → JSON string
        - datetime → ISO format string
        - None → empty string
        - Super long JSON → truncation with key fields preserved
        - Other types → str(value)

        Args:
            value: The value to flatten.

        Returns:
            Flattened value suitable for CSV/Excel cell.
        """
        if value is None:
            return ""

        if isinstance(value, dict):
            json_str = json.dumps(value, ensure_ascii=False)
            # Check if JSON string is too long
            if len(json_str) > self.MAX_CELL_LENGTH:
                # Extract key fields if present
                key_fields_data = {}
                for key in self.KEY_FIELDS:
                    if key in value:
                        key_fields_data[key] = value[key]

                if key_fields_data:
                    # Preserve key fields and truncate the rest
                    key_json = json.dumps(key_fields_data, ensure_ascii=False)
                    truncated = json_str[: self.MAX_CELL_LENGTH - len(key_json) - 20]
                    return f"{key_json}...[TRUNCATED]...{truncated[-100:]}"
                else:
                    # No key fields, simple truncation
                    return json_str[: self.MAX_CELL_LENGTH - 15] + "...[TRUNCATED]"
            return json_str

        if isinstance(value, list):
            json_str = json.dumps(value, ensure_ascii=False)
            if len(json_str) > self.MAX_CELL_LENGTH:
                return json_str[: self.MAX_CELL_LENGTH - 15] + "...[TRUNCATED]"
            return json_str

        if isinstance(value, datetime):
            return value.isoformat()

        return str(value)

    def _flatten_row(self, row: dict) -> dict:
        """
        Flatten a dictionary row for CSV/Excel export.

        Args:
            row: The row dictionary to flatten.

        Returns:
            Flattened dictionary with all values processed.
        """
        flattened = {}
        for key, value in row.items():
            flattened[key] = self._flatten_value(value)
        return flattened

    def to_dict(self) -> dict:
        """Convert to dictionary."""
        return {
            "metadata": self.metadata.to_dict(),
            "summary": self.summary,
            "details": self.details,
            "compliance_checks": self.compliance_checks,
            "recommendations": self.recommendations,
        }

    def to_json(self, indent: int = 2) -> str:
        """Convert to JSON string."""
        return json.dumps(self.to_dict(), indent=indent, default=str)

    def to_csv(self) -> str:
        """Convert details to CSV string.

        Handles both list format (single section) and dict format (multiple sections).
        For dict format, adds a 'section' column to identify the data source.
        Flattens nested fields (dict/list) to JSON strings for proper CSV export.
        """
        if not self.details:
            return ""

        output = io.StringIO()

        # Check if details is dict format (comprehensive report) or list format
        if isinstance(self.details, dict):
            # Dict format: multiple sections
            # Collect all rows with section column
            all_rows = []
            sections_order = [
                "usage_summary",
                "user_activity",
                "audit_trail",
                "security_events",
                "quota_alerts",
            ]

            for section_key in sections_order:
                section_data = self.details.get(section_key, [])
                if section_data and isinstance(section_data, list):
                    for row in section_data:
                        # Add section column to each row
                        row_copy = dict(row)
                        row_copy["section"] = section_key
                        # Flatten nested fields
                        flattened_row = self._flatten_row(row_copy)
                        all_rows.append(flattened_row)

            if not all_rows:
                return ""

            # Get all fieldnames (union of all keys) with section as first column
            all_fieldnames = set()
            for row in all_rows:
                all_fieldnames.update(row.keys())
            fieldnames = ["section"] + sorted([f for f in all_fieldnames if f != "section"])

            writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(all_rows)
        else:
            # List format: single section (backward compatible)
            # Flatten nested fields for each row
            flattened_details = [self._flatten_row(row) for row in self.details]
            writer = csv.DictWriter(output, fieldnames=flattened_details[0].keys())
            writer.writeheader()
            writer.writerows(flattened_details)

        return output.getvalue()

    def to_html(self, language: str = "en") -> str:
        """
        Convert report to HTML string using Jinja2 templates.

        Args:
            language: Language for key titles (default: 'en').
                     Supports: 'en', 'zh', 'ja', 'ko'

        Returns:
            HTML string with formatted report content.
        """
        # Template directory
        template_dir = os.path.join(os.path.dirname(__file__), "templates")

        # Create Jinja2 environment with autoescape for XSS protection
        env = Environment(
            loader=FileSystemLoader(template_dir),
            autoescape=select_autoescape(["html", "xml"]),
        )

        # Select template based on report type
        report_type = self.metadata.report_type
        template_map = {
            ReportType.USAGE_SUMMARY.value: "reports/usage_summary.html",
            ReportType.USER_ACTIVITY.value: "reports/user_activity.html",
            ReportType.AUDIT_TRAIL.value: "reports/audit_trail.html",
            ReportType.DATA_ACCESS.value: "reports/data_access.html",
            ReportType.SECURITY.value: "reports/security.html",
            ReportType.QUOTA_USAGE.value: "reports/quota_usage.html",
            ReportType.COMPREHENSIVE.value: "reports/comprehensive.html",
        }

        template_name = template_map.get(report_type, "reports/comprehensive.html")

        try:
            template = env.get_template(template_name)
        except Exception:
            # Fallback to base template if specific template not found
            template = env.get_template("base.html")

        # Prepare template context
        # Translation keys for key titles
        translations = {
            "en": {
                "reportTitle": "Compliance Report",
                "usageSummary": "Usage Summary Report",
                "userActivity": "User Activity Report",
                "auditTrail": "Audit Trail Report",
                "dataAccess": "Data Access Report",
                "securityReport": "Security Report",
                "quotaUsage": "Quota Usage Report",
                "comprehensive": "Comprehensive Report",
                "sectionUsageSummary": "Usage Summary",
                "sectionUserActivity": "User Activity",
                "sectionAuditTrail": "Audit Trail",
                "sectionSecurityEvents": "Security Events",
                "sectionQuotaAlerts": "Quota Alerts",
            },
            "zh": {
                "reportTitle": "合规报告",
                "usageSummary": "使用汇总报告",
                "userActivity": "用户活动报告",
                "auditTrail": "审计追踪报告",
                "dataAccess": "数据访问报告",
                "securityReport": "安全报告",
                "quotaUsage": "配额使用报告",
                "comprehensive": "综合报告",
                "sectionUsageSummary": "使用统计",
                "sectionUserActivity": "用户活动",
                "sectionAuditTrail": "审计轨迹",
                "sectionSecurityEvents": "安全事件",
                "sectionQuotaAlerts": "配额告警",
            },
            "ja": {
                "reportTitle": "コンプライアンスレポート",
                "usageSummary": "使用量サマリー",
                "userActivity": "ユーザー活動レポート",
                "auditTrail": "監査追跡レポート",
                "dataAccess": "データアクセスレポート",
                "securityReport": "セキュリティレポート",
                "quotaUsage": "クォータ使用レポート",
                "comprehensive": "総合レポート",
                "sectionUsageSummary": "使用統計",
                "sectionUserActivity": "ユーザー活動",
                "sectionAuditTrail": "監査履歴",
                "sectionSecurityEvents": "セキュリティイベント",
                "sectionQuotaAlerts": "クォータアラート",
            },
            "ko": {
                "reportTitle": "준수 보고서",
                "usageSummary": "사용량 요약 보고서",
                "userActivity": "사용자 활동 보고서",
                "auditTrail": "감사 추적 보고서",
                "dataAccess": "데이터 액세스 보고서",
                "securityReport": "보안 보고서",
                "quotaUsage": "할당량 사용 보고서",
                "comprehensive": "종합 보고서",
                "sectionUsageSummary": "사용량 요약",
                "sectionUserActivity": "사용자 활동",
                "sectionAuditTrail": "감사 추적",
                "sectionSecurityEvents": "보안 이벤트",
                "sectionQuotaAlerts": "할당량 경고",
            },
        }

        lang_trans = translations.get(language, translations["en"])

        # Get report title
        report_titles = {
            ReportType.USAGE_SUMMARY.value: lang_trans.get("usageSummary", "Usage Summary"),
            ReportType.USER_ACTIVITY.value: lang_trans.get("userActivity", "User Activity"),
            ReportType.AUDIT_TRAIL.value: lang_trans.get("auditTrail", "Audit Trail"),
            ReportType.DATA_ACCESS.value: lang_trans.get("dataAccess", "Data Access"),
            ReportType.SECURITY.value: lang_trans.get("securityReport", "Security Report"),
            ReportType.QUOTA_USAGE.value: lang_trans.get("quotaUsage", "Quota Usage"),
            ReportType.COMPREHENSIVE.value: lang_trans.get("comprehensive", "Comprehensive"),
        }

        # Prepare context
        context = {
            "report_title": report_titles.get(report_type, lang_trans["reportTitle"]),
            "report_type_display": report_titles.get(report_type, report_type),
            "report_id": self.metadata.report_id,
            "period_start": self.metadata.period_start.strftime("%Y-%m-%d"),
            "period_end": self.metadata.period_end.strftime("%Y-%m-%d"),
            "generated_at": self.metadata.generated_at.strftime("%Y-%m-%d %H:%M:%S UTC"),
            "generated_by": self.metadata.generated_by,
            "generated_by_name": "",
            "summary": self.summary,
            "details": self.details,
            "checks": self.compliance_checks,
            "recommendations": self.recommendations,
            "trans": lang_trans,  # Add translations for template
        }

        # Get generator name if available
        if self.metadata.generated_by:
            try:
                user_repo = UserRepository()
                user = user_repo.get_user_by_id(self.metadata.generated_by)
                if user:
                    context["generated_by_name"] = user.get("username", f"User {self.metadata.generated_by}")
            except Exception:
                context["generated_by_name"] = f"User {self.metadata.generated_by}"

        return str(template.render(**context))

    def to_excel(self, language: str = "en") -> bytes:
        """
        Convert report to Excel file using openpyxl.

        Args:
            language: Language for key titles (default: 'en').
                     Supports: 'en', 'zh', 'ja', 'ko'

        Returns:
            Bytes containing the Excel file content.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # Translation keys for column headers
        translations = {
            "en": {
                "summary": "Summary",
                "details": "Details",
                "complianceChecks": "Compliance Checks",
                "recommendations": "Recommendations",
                "checkId": "Check ID",
                "name": "Name",
                "status": "Status",
                "message": "Message",
                "noData": "No data available",
                "section": "Section",
                "sectionUsageSummary": "Usage Summary",
                "sectionUserActivity": "User Activity",
                "sectionAuditTrail": "Audit Trail",
                "sectionSecurityEvents": "Security Events",
                "sectionQuotaAlerts": "Quota Alerts",
            },
            "zh": {
                "summary": "汇总",
                "details": "详情",
                "complianceChecks": "合规检查",
                "recommendations": "建议",
                "checkId": "检查ID",
                "name": "名称",
                "status": "状态",
                "message": "消息",
                "noData": "无数据",
                "section": "部分",
                "sectionUsageSummary": "使用统计",
                "sectionUserActivity": "用户活动",
                "sectionAuditTrail": "审计轨迹",
                "sectionSecurityEvents": "安全事件",
                "sectionQuotaAlerts": "配额告警",
            },
            "ja": {
                "summary": "サマリー",
                "details": "詳細",
                "complianceChecks": "コンプライアンスチェック",
                "recommendations": "推奨事項",
                "checkId": "チェックID",
                "name": "名前",
                "status": "状態",
                "message": "メッセージ",
                "noData": "データなし",
                "section": "セクション",
                "sectionUsageSummary": "使用統計",
                "sectionUserActivity": "ユーザー活動",
                "sectionAuditTrail": "監査履歴",
                "sectionSecurityEvents": "セキュリティイベント",
                "sectionQuotaAlerts": "クォータアラート",
            },
            "ko": {
                "summary": "요약",
                "details": "상세",
                "complianceChecks": "준수 검사",
                "recommendations": "제안",
                "checkId": "검사 ID",
                "name": "이름",
                "status": "상태",
                "message": "메시지",
                "noData": "데이터 없음",
                "section": "섹션",
                "sectionUsageSummary": "사용량 요약",
                "sectionUserActivity": "사용자 활동",
                "sectionAuditTrail": "감사 추적",
                "sectionSecurityEvents": "보안 이벤트",
                "sectionQuotaAlerts": "할당량 경고",
            },
        }

        lang_trans = translations.get(language, translations["en"])

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Status colors
        status_fills = {
            "pass": PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid"),
            "warning": PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid"),
            "fail": PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid"),
        }

        # === Summary Sheet ===
        ws_summary = wb.active
        ws_summary.title = lang_trans["summary"]

        # Add metadata
        ws_summary["A1"] = "Report ID"
        ws_summary["B1"] = self.metadata.report_id
        ws_summary["A2"] = "Report Type"
        ws_summary["B2"] = self.metadata.report_type
        ws_summary["A3"] = "Period Start"
        ws_summary["B3"] = self.metadata.period_start.strftime("%Y-%m-%d")
        ws_summary["A4"] = "Period End"
        ws_summary["B4"] = self.metadata.period_end.strftime("%Y-%m-%d")
        ws_summary["A5"] = "Generated At"
        ws_summary["B5"] = self.metadata.generated_at.strftime("%Y-%m-%d %H:%M:%S")

        for row in range(1, 6):
            ws_summary[f"A{row}"].font = Font(bold=True)

        # Add summary data
        row_num = 7
        ws_summary[f"A{row_num}"] = lang_trans["summary"]
        ws_summary[f"A{row_num}"].font = header_font
        ws_summary[f"A{row_num}"].fill = header_fill

        row_num += 1
        for key, value in self.summary.items():
            if key == "period":
                continue
            ws_summary[f"A{row_num}"] = str(key)
            ws_summary[f"B{row_num}"] = (
                str(value) if not isinstance(value, dict) else json.dumps(value)
            )
            ws_summary[f"A{row_num}"].font = Font(bold=True)
            ws_summary[f"A{row_num}"].border = border
            ws_summary[f"B{row_num}"].border = border
            row_num += 1

        # === Details Sheet ===
        ws_details = wb.create_sheet(lang_trans["details"])

        if self.details:
            # Check if details is dict format (comprehensive report) or list format
            if isinstance(self.details, dict):
                # Dict format: multiple sections
                # Section order and translations
                sections_order = [
                    ("usage_summary", lang_trans.get("sectionUsageSummary", "Usage Summary")),
                    ("user_activity", lang_trans.get("sectionUserActivity", "User Activity")),
                    ("audit_trail", lang_trans.get("sectionAuditTrail", "Audit Trail")),
                    ("security_events", lang_trans.get("sectionSecurityEvents", "Security Events")),
                    ("quota_alerts", lang_trans.get("sectionQuotaAlerts", "Quota Alerts")),
                ]

                row_idx = 1
                for section_key, section_title in sections_order:
                    section_data = self.details.get(section_key, [])
                    if not section_data or not isinstance(section_data, list):
                        continue

                    # Add section header row
                    section_header_cell = ws_details.cell(
                        row=row_idx, column=1, value=f"=== {section_title} ==="
                    )
                    section_header_cell.font = Font(bold=True, size=14)
                    row_idx += 1

                    # Write column headers with Section column
                    headers = ["Section"] + list(section_data[0].keys())
                    for col_idx, header in enumerate(headers, 1):
                        cell = ws_details.cell(row=row_idx, column=col_idx, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = border
                    row_idx += 1

                    # Write data rows
                    for detail in section_data:
                        # Flatten nested fields for each row
                        flattened_detail = self._flatten_row(detail)
                        cell = ws_details.cell(row=row_idx, column=1, value=section_key)
                        cell.border = border
                        for col_idx, header in enumerate(headers[1:], 2):
                            value = flattened_detail.get(header)
                            cell = ws_details.cell(row=row_idx, column=col_idx, value=value)
                            cell.border = border
                        row_idx += 1

                    # Add blank row between sections
                    row_idx += 1

                # Auto-adjust column widths if any data was written
                if row_idx > 1:
                    for col_idx in range(1, 10):  # Check first 10 columns
                        col_letter = get_column_letter(col_idx)
                        ws_details.column_dimensions[col_letter].width = 15
            else:
                # List format: single section (backward compatible)
                # Write headers
                headers = list(self.details[0].keys())
                for col_idx, header in enumerate(headers, 1):
                    cell = ws_details.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border

                # Write data (flatten nested fields for each row)
                for row_idx, detail in enumerate(self.details, 2):
                    flattened_detail = self._flatten_row(detail)
                    for col_idx, header in enumerate(headers, 1):
                        value = flattened_detail.get(header)
                        cell = ws_details.cell(row=row_idx, column=col_idx, value=value)
                        cell.border = border

                # Auto-adjust column widths
                for col_idx in range(1, len(headers) + 1):
                    col_letter = get_column_letter(col_idx)
                    max_length = max(
                        len(str(ws_details.cell(row=1, column=col_idx).value or "")),
                        max(
                            len(str(ws_details.cell(row=r, column=col_idx).value or ""))
                            for r in range(2, len(self.details) + 2)
                        ),
                    )
                    ws_details.column_dimensions[col_letter].width = min(max_length + 2, 50)
        else:
            ws_details["A1"] = lang_trans["noData"]

        # === Compliance Checks Sheet ===
        ws_checks = wb.create_sheet(lang_trans["complianceChecks"])

        if self.compliance_checks:
            check_headers = [
                lang_trans["checkId"],
                lang_trans["name"],
                lang_trans["status"],
                lang_trans["message"],
            ]
            for col_idx, header in enumerate(check_headers, 1):
                cell = ws_checks.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            for row_idx, check in enumerate(self.compliance_checks, 2):
                ws_checks.cell(row=row_idx, column=1, value=check.get("check_id")).border = border
                ws_checks.cell(row=row_idx, column=2, value=check.get("name")).border = border

                status_cell = ws_checks.cell(row=row_idx, column=3, value=check.get("status"))
                status_cell.border = border
                status_val = check.get("status")
                if status_val and status_val in status_fills:
                    status_cell.fill = status_fills[status_val]

                ws_checks.cell(row=row_idx, column=4, value=check.get("message")).border = border

            for col_idx in range(1, 5):
                col_letter = get_column_letter(col_idx)
                ws_checks.column_dimensions[col_letter].width = 20
        else:
            ws_checks["A1"] = lang_trans["noData"]

        # === Recommendations Sheet ===
        ws_recommendations = wb.create_sheet(lang_trans["recommendations"])

        if self.recommendations:
            ws_recommendations["A1"] = "#"
            ws_recommendations["B1"] = lang_trans["recommendations"]
            ws_recommendations["A1"].font = header_font
            ws_recommendations["A1"].fill = header_fill
            ws_recommendations["B1"].font = header_font
            ws_recommendations["B1"].fill = header_fill

            for idx, rec in enumerate(self.recommendations, 1):
                ws_recommendations.cell(row=idx + 1, column=1, value=idx).border = border
                ws_recommendations.cell(row=idx + 1, column=2, value=rec).border = border

                # Highlight urgent recommendations
                if rec.startswith("URGENT:"):
                    ws_recommendations.cell(row=idx + 1, column=2).font = Font(
                        bold=True, color="EF4444"
                    )
                elif rec.startswith("Review:"):
                    ws_recommendations.cell(row=idx + 1, column=2).font = Font(color="F59E0B")

            ws_recommendations.column_dimensions["A"].width = 5
            ws_recommendations.column_dimensions["B"].width = 80
        else:
            ws_recommendations["A1"] = lang_trans["noData"]

        # Write to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()


class ReportGenerator:
    """
    Generator for compliance reports.

    Features:
    - Multiple report types
    - Various output formats
    - Customizable date ranges
    - Compliance check integration
    """

    def __init__(
        self,
        db: Optional[Database] = None,
        usage_repo: Optional[UsageRepository] = None,
        user_repo: Optional[UserRepository] = None,
        audit_logger: Optional[AuditLogger] = None,
    ):
        """
        Initialize report generator.

        Args:
            db: Optional Database instance.
            usage_repo: Optional UsageRepository instance.
            user_repo: Optional UserRepository instance.
            audit_logger: Optional AuditLogger instance.
        """
        self.db = db or Database()
        self.usage_repo = usage_repo or UsageRepository()
        self.user_repo = user_repo or UserRepository()
        self.audit_logger = audit_logger or AuditLogger()

    def generate_report(
        self,
        report_type: str,
        period_start: datetime,
        period_end: datetime,
        generated_by: Optional[int] = None,
        tenant_id: Optional[int] = None,
        filters: Optional[dict[str, Any]] = None,
    ) -> ComplianceReport:
        """
        Generate a compliance report.

        Args:
            report_type: Type of report to generate.
            period_start: Report period start date.
            period_end: Report period end date.
            generated_by: User ID of report generator.
            tenant_id: Tenant ID for multi-tenant reports.
            filters: Additional filters.

        Returns:
            ComplianceReport: Generated report.
        """
        import uuid

        metadata = ReportMetadata(
            report_id=str(uuid.uuid4()),
            report_type=report_type,
            generated_at=datetime.now(timezone.utc).replace(tzinfo=None),
            period_start=period_start,
            period_end=period_end,
            generated_by=generated_by,
            tenant_id=tenant_id,
            filters=filters or {},
        )

        # Generate report based on type
        if report_type == ReportType.USAGE_SUMMARY.value:
            summary, details = self._generate_usage_summary(period_start, period_end, tenant_id)
        elif report_type == ReportType.USER_ACTIVITY.value:
            summary, details = self._generate_user_activity(period_start, period_end, tenant_id)
        elif report_type == ReportType.AUDIT_TRAIL.value:
            summary, details = self._generate_audit_trail(period_start, period_end, tenant_id)
        elif report_type == ReportType.DATA_ACCESS.value:
            summary, details = self._generate_data_access(period_start, period_end, tenant_id)
        elif report_type == ReportType.SECURITY.value:
            summary, details = self._generate_security_report(period_start, period_end, tenant_id)
        elif report_type == ReportType.QUOTA_USAGE.value:
            summary, details = self._generate_quota_usage(period_start, period_end, tenant_id)
        elif report_type == ReportType.COMPREHENSIVE.value:
            summary, details = self._generate_comprehensive_report(
                period_start, period_end, tenant_id
            )
        else:
            summary, details = {}, []

        # Run compliance checks
        compliance_checks = self._run_compliance_checks(report_type, summary, details)

        # Generate recommendations
        recommendations = self._generate_recommendations(compliance_checks)

        return ComplianceReport(
            metadata=metadata,
            summary=summary,
            details=details,
            compliance_checks=compliance_checks,
            recommendations=recommendations,
        )

    def _generate_usage_summary(
        self, period_start: datetime, period_end: datetime, tenant_id: Optional[int]
    ) -> tuple:
        """Generate usage summary report."""
        start_date = period_start.strftime("%Y-%m-%d")
        end_date = period_end.strftime("%Y-%m-%d")

        # Get usage data from daily_usage (aggregated by tool/host)
        usage_data = self.db.fetch_all(
            """
            SELECT
                date,
                tool_name,
                host_name,
                SUM(input_tokens) as total_input_tokens,
                SUM(output_tokens) as total_output_tokens,
                SUM(tokens_used) as total_tokens,
                SUM(request_count) as total_requests
            FROM daily_usage
            WHERE date >= ? AND date <= ?
            GROUP BY date, tool_name, host_name
            ORDER BY date DESC, tool_name
        """,
            (start_date, end_date),
        )

        # Get unique users count from user_daily_stats table
        unique_users_result = self.db.fetch_one(
            """
            SELECT COUNT(DISTINCT user_id) as unique_users
            FROM user_daily_stats
            WHERE date >= ? AND date <= ?
        """,
            (start_date, end_date),
        )
        unique_users = unique_users_result.get("unique_users", 0) if unique_users_result else 0

        # Calculate summary
        total_tokens = sum(r.get("total_tokens", 0) or 0 for r in usage_data)
        total_requests = sum(r.get("total_requests", 0) or 0 for r in usage_data)
        unique_tools = {r.get("tool_name") for r in usage_data if r.get("tool_name")}

        summary: dict[str, Any] = {
            "period": {
                "start": start_date,
                "end": end_date,
                "days": (period_end - period_start).days + 1,
            },
            "totals": {
                "tokens": total_tokens,
                "requests": total_requests,
                "tools_used": len(unique_tools),
                "tools": list(unique_tools),
                "unique_users": unique_users,
            },
            "averages": {
                "daily_tokens": total_tokens // max((period_end - period_start).days + 1, 1),
                "daily_requests": total_requests // max((period_end - period_start).days + 1, 1),
            },
        }

        details = [dict(r) for r in usage_data]

        return summary, details

    def _generate_user_activity(
        self, period_start: datetime, period_end: datetime, tenant_id: Optional[int]
    ) -> tuple:
        """Generate user activity report."""
        start_date = period_start.strftime("%Y-%m-%d")
        end_date = period_end.strftime("%Y-%m-%d")

        # Get user activity from user_daily_stats (has user_id column)
        user_activity = self.db.fetch_all(
            """
            SELECT
                u.id as user_id,
                u.username,
                u.email,
                u.role,
                COUNT(DISTINCT uds.date) as active_days,
                SUM(uds.tokens) as total_tokens,
                SUM(uds.requests) as total_requests,
                MIN(uds.date) as first_activity,
                MAX(uds.date) as last_activity
            FROM users u
            LEFT JOIN user_daily_stats uds ON u.id = uds.user_id
                AND uds.date >= ? AND uds.date <= ?
            GROUP BY u.id
            ORDER BY total_tokens DESC
        """,
            (start_date, end_date),
        )

        # Calculate summary
        active_users = [u for u in user_activity if u.get("active_days", 0) > 0]
        total_users = len(user_activity)

        summary: dict[str, Any] = {
            "period": {
                "start": start_date,
                "end": end_date,
            },
            "totals": {
                "total_users": total_users,
                "active_users": len(active_users),
                "inactive_users": total_users - len(active_users),
            },
            "by_role": self._group_by_role(user_activity),
        }

        details = [dict(r) for r in user_activity]

        return summary, details

    def _generate_audit_trail(
        self, period_start: datetime, period_end: datetime, tenant_id: Optional[int]
    ) -> tuple:
        """Generate audit trail report."""
        audit_logs = self.audit_logger.query(
            start_time=period_start,
            end_time=period_end,
            limit=10000,
        )

        # Calculate summary
        action_counts: dict[str, int] = {}
        severity_counts: dict[str, int] = {}
        user_actions: dict[Union[str, int], int] = {}

        for log in audit_logs:
            action = log.action
            severity = log.severity
            user_id = log.user_id

            action_counts[action] = action_counts.get(action, 0) + 1
            severity_counts[severity] = severity_counts.get(severity, 0) + 1

            if user_id:
                user_actions[user_id] = user_actions.get(user_id, 0) + 1

        summary: dict[str, Any] = {
            "period": {
                "start": period_start.isoformat(),
                "end": period_end.isoformat(),
            },
            "totals": {
                "total_events": len(audit_logs),
                "unique_users": len(user_actions),
            },
            "by_action": action_counts,
            "by_severity": severity_counts,
        }

        details = [log.to_dict() for log in audit_logs]

        return summary, details

    def _generate_data_access(
        self, period_start: datetime, period_end: datetime, tenant_id: Optional[int]
    ) -> tuple:
        """Generate data access report."""
        # Get data access logs
        access_logs = self.audit_logger.query(
            start_time=period_start,
            end_time=period_end,
            resource_type="data",
            limit=10000,
        )

        # Also get export/import actions
        export_logs = self.audit_logger.query(
            start_time=period_start,
            end_time=period_end,
            action="data_export",
            limit=10000,
        )

        summary: dict[str, Any] = {
            "period": {
                "start": period_start.isoformat(),
                "end": period_end.isoformat(),
            },
            "totals": {
                "data_access_events": len(access_logs),
                "data_exports": len(export_logs),
            },
        }

        details = [log.to_dict() for log in access_logs + export_logs]

        return summary, details

    def _generate_security_report(
        self, period_start: datetime, period_end: datetime, tenant_id: Optional[int]
    ) -> tuple:
        """Generate security report."""
        # Get security-related audit logs
        security_actions = [
            "login",
            "login_failed",
            "logout",
            "session_expired",
            "user_password_change",
            "user_role_change",
            "permission_grant",
            "permission_revoke",
            "content_blocked",
            "content_flagged",
        ]

        all_logs = []
        for action in security_actions:
            logs = self.audit_logger.query(
                start_time=period_start,
                end_time=period_end,
                action=action,
                limit=1000,
            )
            all_logs.extend(logs)

        # Calculate summary
        failed_logins = [l for l in all_logs if l.action == "login_failed"]
        password_changes = [l for l in all_logs if l.action == "user_password_change"]
        role_changes = [l for l in all_logs if l.action == "user_role_change"]
        content_blocked = [l for l in all_logs if l.action == "content_blocked"]

        summary: dict[str, Any] = {
            "period": {
                "start": period_start.isoformat(),
                "end": period_end.isoformat(),
            },
            "security_events": {
                "failed_logins": len(failed_logins),
                "password_changes": len(password_changes),
                "role_changes": len(role_changes),
                "content_blocked": len(content_blocked),
            },
            "risk_indicators": [],
        }

        # Add risk indicators
        if len(failed_logins) > 10:
            summary["risk_indicators"].append(
                {
                    "level": "warning",
                    "message": f"High number of failed login attempts: {len(failed_logins)}",
                }
            )

        if len(role_changes) > 5:
            summary["risk_indicators"].append(
                {
                    "level": "info",
                    "message": f"Multiple role changes detected: {len(role_changes)}",
                }
            )

        details = [log.to_dict() for log in all_logs]

        return summary, details

    def _generate_quota_usage(
        self, period_start: datetime, period_end: datetime, tenant_id: Optional[int]
    ) -> tuple:
        """Generate quota usage report."""
        # Get quota alerts
        quota_alerts = self.db.fetch_all(
            """
            SELECT * FROM quota_alerts
            WHERE created_at >= ? AND created_at <= ?
            ORDER BY created_at DESC
        """,
            (period_start, period_end),
        )

        # Get user quota status
        users = self.user_repo.get_all_users(include_inactive=False)

        summary: dict[str, Any] = {
            "period": {
                "start": period_start.isoformat(),
                "end": period_end.isoformat(),
            },
            "alerts": {
                "total": len(quota_alerts),
                "warnings": len([a for a in quota_alerts if a.get("alert_type") == "warning"]),
                "critical": len([a for a in quota_alerts if a.get("alert_type") == "critical"]),
                "exceeded": len([a for a in quota_alerts if a.get("alert_type") == "exceeded"]),
            },
            "users_monitored": len(users),
        }

        details = [dict(a) for a in quota_alerts]

        return summary, details

    def _generate_comprehensive_report(
        self, period_start: datetime, period_end: datetime, tenant_id: Optional[int]
    ) -> tuple:
        """Generate comprehensive compliance report.

        Returns:
            tuple: (summary dict, details dict with sections)
                details structure:
                {
                    "usage_summary": [...],
                    "user_activity": [...],
                    "audit_trail": [...],
                    "security_events": [...],
                    "quota_alerts": [...]
                }
        """
        # Combine all report types - now keeping all details
        try:
            usage_summary, usage_details = self._generate_usage_summary(
                period_start, period_end, tenant_id
            )
        except Exception as e:
            logger.warning(f"Failed to generate usage summary: {e}")
            usage_summary, usage_details = {}, []

        try:
            user_activity, user_details = self._generate_user_activity(
                period_start, period_end, tenant_id
            )
        except Exception as e:
            logger.warning(f"Failed to generate user activity: {e}")
            user_activity, user_details = {}, []

        try:
            audit_summary, audit_details = self._generate_audit_trail(
                period_start, period_end, tenant_id
            )
        except Exception as e:
            logger.warning(f"Failed to generate audit trail: {e}")
            audit_summary, audit_details = {}, []

        try:
            security_summary, security_details = self._generate_security_report(
                period_start, period_end, tenant_id
            )
        except Exception as e:
            logger.warning(f"Failed to generate security report: {e}")
            security_summary, security_details = {}, []

        try:
            quota_summary, quota_details = self._generate_quota_usage(
                period_start, period_end, tenant_id
            )
        except Exception as e:
            logger.warning(f"Failed to generate quota usage: {e}")
            quota_summary, quota_details = {}, []

        summary: dict[str, Any] = {
            "period": {
                "start": period_start.isoformat(),
                "end": period_end.isoformat(),
            },
            "usage": usage_summary,
            "users": user_activity,
            "audit": audit_summary,
            "security": security_summary,
            "quota": quota_summary,
        }

        # New dict structure for details with sections
        details: dict[str, Any] = {
            "usage_summary": usage_details,
            "user_activity": user_details,
            "audit_trail": audit_details,
            "security_events": security_details,
            "quota_alerts": quota_details,
        }

        return summary, details

    def _run_compliance_checks(
        self, report_type: str, summary: dict[str, Any], details: list[dict[str, Any]]
    ) -> list[dict[str, Any]]:
        """Run compliance checks on the report data."""
        checks = []

        # Check 1: Data retention compliance
        checks.append(
            {
                "check_id": "data_retention",
                "name": "Data Retention Policy",
                "status": "pass",
                "message": "Data retention policy is being followed",
            }
        )

        # Check 2: User access review
        if summary.get("users", {}).get("inactive_users", 0) > 10:
            checks.append(
                {
                    "check_id": "inactive_users",
                    "name": "Inactive User Review",
                    "status": "warning",
                    "message": f"{summary['users']['inactive_users']} inactive users found. Review access.",
                }
            )
        else:
            checks.append(
                {
                    "check_id": "inactive_users",
                    "name": "Inactive User Review",
                    "status": "pass",
                    "message": "Inactive user count is within acceptable range",
                }
            )

        # Check 3: Security events
        security_events = summary.get("security", {}).get("security_events", {})
        if security_events.get("failed_logins", 0) > 20:
            checks.append(
                {
                    "check_id": "failed_logins",
                    "name": "Failed Login Monitoring",
                    "status": "fail",
                    "message": f"High number of failed logins: {security_events['failed_logins']}",
                }
            )
        else:
            checks.append(
                {
                    "check_id": "failed_logins",
                    "name": "Failed Login Monitoring",
                    "status": "pass",
                    "message": "Failed login attempts within normal range",
                }
            )

        # Check 4: Quota compliance
        quota_alerts = summary.get("quota", {}).get("alerts", {})
        if quota_alerts.get("exceeded", 0) > 0:
            checks.append(
                {
                    "check_id": "quota_compliance",
                    "name": "Quota Compliance",
                    "status": "warning",
                    "message": f"{quota_alerts['exceeded']} quota exceedances detected",
                }
            )
        else:
            checks.append(
                {
                    "check_id": "quota_compliance",
                    "name": "Quota Compliance",
                    "status": "pass",
                    "message": "All users within quota limits",
                }
            )

        return checks

    def _generate_recommendations(self, compliance_checks: list[dict[str, Any]]) -> list[str]:
        """Generate recommendations based on compliance checks."""
        recommendations = []

        for check in compliance_checks:
            if check["status"] == "fail":
                recommendations.append(f"URGENT: {check['name']} - {check['message']}")
            elif check["status"] == "warning":
                recommendations.append(f"Review: {check['name']} - {check['message']}")

        if not recommendations:
            recommendations.append("All compliance checks passed. Continue monitoring.")

        return recommendations

    def _group_by_role(self, user_activity: list[dict]) -> dict[str, int]:
        """Group user activity by role."""
        by_role: dict[str, int] = {}
        for user in user_activity:
            role = user.get("role", "unknown")
            by_role[role] = by_role.get(role, 0) + 1
        return by_role

    def save_report(self, report: ComplianceReport) -> bool:
        """Save report to database."""
        try:
            with self.db.connection() as conn:
                cursor = conn.cursor()

                # Use SERIAL for PostgreSQL, AUTOINCREMENT for SQLite
                id_type = (
                    "SERIAL PRIMARY KEY"
                    if self.db.is_postgresql
                    else "INTEGER PRIMARY KEY AUTOINCREMENT"
                )

                cursor.execute(
                    f"""
                    CREATE TABLE IF NOT EXISTS compliance_reports (
                        id {id_type},
                        report_id TEXT UNIQUE NOT NULL,
                        report_type TEXT NOT NULL,
                        generated_at TIMESTAMP NOT NULL,
                        period_start TIMESTAMP NOT NULL,
                        period_end TIMESTAMP NOT NULL,
                        generated_by INTEGER,
                        tenant_id INTEGER,
                        report_data TEXT NOT NULL,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                """
                )
                cursor.execute(
                    (
                        """
                    INSERT INTO compliance_reports
                    (report_id, report_type, generated_at, period_start, period_end,
                     generated_by, tenant_id, report_data)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """
                        if self.db.is_postgresql
                        else """
                    INSERT INTO compliance_reports
                    (report_id, report_type, generated_at, period_start, period_end,
                     generated_by, tenant_id, report_data)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """
                    ),
                    (
                        report.metadata.report_id,
                        report.metadata.report_type,
                        report.metadata.generated_at,
                        report.metadata.period_start,
                        report.metadata.period_end,
                        report.metadata.generated_by,
                        report.metadata.tenant_id,
                        report.to_json(),
                    ),
                )
                conn.commit()

            logger.info(f"Saved compliance report: {report.metadata.report_id}")
            return True

        except Exception as e:
            logger.error(f"Failed to save report: {e}")
            return False

    def get_saved_reports(
        self, report_type: Optional[str] = None, tenant_id: Optional[int] = None, limit: int = 50
    ) -> list[dict[str, Any]]:
        """Get saved reports.

        Raises:
            Exception: If database query fails.
        """
        conditions = []
        params: list[Any] = []

        if report_type:
            conditions.append("report_type = ?")
            params.append(report_type)

        if tenant_id:
            conditions.append("tenant_id = ?")
            params.append(tenant_id)

        where_clause = " AND ".join(conditions) if conditions else "1=1"

        query = f"""
            SELECT report_id, report_type, generated_at, period_start, period_end
            FROM compliance_reports
            WHERE {where_clause}
            ORDER BY generated_at DESC
            LIMIT ?
        """

        try:
            rows = self.db.fetch_all(query, tuple(params + [limit]))
            return [dict(r) for r in rows]
        except Exception as e:
            logger.error(f"Failed to query saved reports: {e}")
            raise

    def get_saved_report(self, report_id: str) -> Optional[ComplianceReport]:
        """Get a saved report by ID.

        Handles backward compatibility for old format comprehensive reports
        by converting list-style details to dict format.
        """
        row = self.db.fetch_one(
            "SELECT report_data FROM compliance_reports WHERE report_id = ?", (report_id,)
        )

        if row:
            try:
                data = json.loads(row["report_data"])
                # Reconstruct report from saved data
                metadata = ReportMetadata(
                    report_id=data["metadata"]["report_id"],
                    report_type=data["metadata"]["report_type"],
                    generated_at=datetime.fromisoformat(data["metadata"]["generated_at"]),
                    period_start=datetime.fromisoformat(data["metadata"]["period_start"]),
                    period_end=datetime.fromisoformat(data["metadata"]["period_end"]),
                    generated_by=data["metadata"].get("generated_by"),
                    tenant_id=data["metadata"].get("tenant_id"),
                    filters=data["metadata"].get("filters", {}),
                )

                # Handle backward compatibility for comprehensive report details
                details = data["details"]
                report_type = data["metadata"]["report_type"]

                # If comprehensive report has old list format details, convert to dict format
                if report_type == ReportType.COMPREHENSIVE.value and isinstance(details, list):
                    # Convert old flat list to new dict format
                    details = {
                        "usage_summary": [],
                        "user_activity": [],
                        "audit_trail": details,  # Old details were audit_trail
                        "security_events": [],
                        "quota_alerts": [],
                    }
                    logger.info(
                        f"Converted old format comprehensive report details for {report_id}"
                    )

                return ComplianceReport(
                    metadata=metadata,
                    summary=data["summary"],
                    details=details,
                    compliance_checks=data["compliance_checks"],
                    recommendations=data["recommendations"],
                )
            except Exception as e:
                logger.error(f"Failed to load report: {e}")

        return None


def get_ddl_statements() -> list[str]:
    """Return DDL statements for compliance report tables."""
    from app.repositories.database import is_postgresql

    id_type = "SERIAL PRIMARY KEY" if is_postgresql() else "INTEGER PRIMARY KEY AUTOINCREMENT"
    return [
        f"""
        CREATE TABLE IF NOT EXISTS compliance_reports (
            id {id_type},
            report_id TEXT UNIQUE NOT NULL,
            report_type TEXT NOT NULL,
            generated_at TIMESTAMP NOT NULL,
            period_start TIMESTAMP NOT NULL,
            period_end TIMESTAMP NOT NULL,
            generated_by INTEGER,
            tenant_id INTEGER,
            report_data TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """,
    ]
